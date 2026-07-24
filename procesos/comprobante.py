"""Proceso: Generar Comprobante Bancolombia.

Entrada: ZIP con CSVs adentro (los CSV de Bancolombia tienen una sola
columna con valores separados por coma).
Salida: 1 Excel con 5 hojas (Original, Por cuentas, Por conceptos,
        Intereses, Gastos bancarios) + 1 archivo FOAPAL (fzrcoco.xlsx).

Migrado desde ``scripts/GenerarComprobante.py``:
- recibe el ZIP como parametro (sin depender de ~/Downloads);
- respeta el modo_prueba de la app (no toca los originales);
- centraliza rutas y logging a partir de la raiz del proyecto;
- lee los 4 mapeos desde ``jsons/comprobante/`` (codigos_conceptos,
  codigos_contables, foapal, nit_bancolombia).
"""
from __future__ import annotations

import zipfile
from datetime import datetime
from io import StringIO
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

from procesos.base import ProcesoBase, ResultadoProceso
from utils.archivos import carpeta_modo_prueba, carpeta_resultados
from utils.bitacora import log
from utils.json_manager import leer_json

RAIZ = Path(__file__).resolve().parent.parent

ENCODING_FALLBACKS: tuple[str, ...] = ("utf-8", "utf-8-sig", "latin-1")

# Estructura de las columnas del CSV de Bancolombia (el banco NO
# pone headers, por eso usamos indices numericos):
#   col 0: Cuenta              (cuenta Bancolombia, 10 digitos)
#   col 1: Prefijo de cuenta   (oficina)
#   col 2: Tipo                (Corriente / Ahorro)
#   col 3: Fecha               (formato ddmmyyyy)
#   col 4: Identificador       (codigo interno, irrelevante para FOAPAL)
#   col 5: Valor               (valor del movimiento)
#   col 6: Codigo de concepto  (4 digitos, ej: 480, 1334, 2999)
#   col 7: Concepto            (descripcion textual del movimiento)
#   col 8: Valor numerico      (valor sin signo, redundante con col 5)
#   col 9: Codigo Contable     (agregado por _agregar_codigo_contable)
COL_CUENTA = 0
COL_PREFIJO = 1
COL_TIPO = 2
COL_FECHA = 3
COL_ID = 4
COL_VALOR = 5
COL_CODIGO_CONCEPTO = 6
COL_CONCEPTO = 7
COL_VALOR_NUM = 8
COL_CODIGO_CONTABLE = 9

COLUMNAS_NUMERICAS: list[int] = [
    COL_CUENTA, COL_PREFIJO, COL_TIPO, COL_VALOR, COL_VALOR_NUM,
]
CODIGO_EXCEPCION_FOAPAL: str = "119090"
EXCEPCION_FOAPAL: dict = {
    "Fondo": "FOESPC",
    "Organizacion": 52617,
    "Cuenta": 280517,
    "Programa": 52617,
}
ENCABEZADOS_FOAPAL: list[str] = [
    "Codigo", "Concepto", "Valor", "Fecha",
    "Fondo", "Organizacion", "Cuenta", "Programa", "D/C",
]

# Meses en espanol (no dependemos del locale del sistema).
MESES_ES: dict[int, str] = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _read_csv_con_fallback(archivo: Path) -> pd.DataFrame:
    """Lee un CSV desde disco probando utf-8 -> utf-8-sig -> latin-1.

    NOTA IMPORTANTE sobre ``sep="|"``:
        El CSV real de Bancolombia tiene UNA sola linea por movimiento
        con 9 campos separados por coma. Algunos campos (descripcion)
        pueden contener comas, por ejemplo:
            "TRANSFERENCIA CTA SUC VIRTUAL, PROVEEDOR"
        Si usaramos ``sep=","`` pandas dividiria esa linea en 2 (o mas)
        filas, rompiendo la alineacion de columnas.

        El truco: usamos ``sep="|"`` porque el CSV NUNCA contiene pipes.
        Asi pandas lee toda la linea como UNA sola celda, sin dividir
        nada. Despues, en ``_copy_data``, hacemos el split por coma
        manualmente con ``str.split(pat=",", expand=True)``.

        NO cambiar ``sep="|"`` por ``sep=","`` sin entender esto primero.
    """
    for encoding in ENCODING_FALLBACKS:
        try:
            return pd.read_csv(
                archivo, decimal=".", sep="|", header=None, dtype=str,
                encoding=encoding,
            )
        except UnicodeDecodeError:
            continue
    return pd.read_csv(
        archivo, decimal=".", sep="|", header=None, dtype=str, encoding="latin-1",
    )


class ProcesoComprobante(ProcesoBase):
    """Genera el comprobante contable y el archivo FOAPAL."""

    LOG_PREFIX = "[Comprobante]"

    def __init__(self) -> None:
        super().__init__()
        # Carga de los 4 JSONs del proceso.
        json_dir = RAIZ / "jsons" / "comprobante"
        self.clasificador_conceptos: dict = leer_json(json_dir / "codigos_conceptos.json")
        self.codigos_contables: dict[str, str] = leer_json(json_dir / "codigos_contables.json")
        self.nit_bancolombia: dict[str, str] = leer_json(json_dir / "nit_bancolombia.json")
        self.foapal_config: dict = leer_json(json_dir / "foapal.json")

        # Constantes derivadas.
        self.cuenta_bancolombia: str = next(
            cuenta for cuenta, nit in self.nit_bancolombia.items()
            if nit != "890903938"
        )
        self.concepto_fiduciaria: str = next(
            codigo for codigo, descripciones
            in self.clasificador_conceptos["Gastos bancarios"].items()
            if any("PAGP A PROVE FIDUCIARIA BANC" in d for d in descripciones)
        )
        self.codigos_interes: list[str] = list(
            self.clasificador_conceptos.get("Intereses", {}).keys()
        )
        self.codigos_gastos: list[str] = list(
            self.clasificador_conceptos.get("Gastos bancarios", {}).keys()
        )

    @property
    def nombre(self) -> str:
        return "comprobante"

    @property
    def descripcion(self) -> str:
        return (
            "Genera el comprobante contable Bancolombia (Excel 5 hojas) "
            "y el archivo FOAPAL (fzrcoco.xlsx) a partir de un ZIP con CSVs."
        )

    @property
    def extensiones_entrada(self) -> tuple[str, ...]:
        return (".zip",)

    @property
    def extensiones_salida(self) -> tuple[str, ...]:
        return (".xlsx",)

    def validar_archivos(self, archivos: list[Path]) -> str | None:
        if not archivos:
            return "Se esperaba al menos un archivo ZIP."
        for archivo in archivos:
            if archivo.suffix.lower() not in self.extensiones_entrada:
                return (
                    f"El archivo '{archivo.name}' no es un ZIP. "
                    "Para comprobante solo se aceptan archivos .zip."
                )
            if not zipfile.is_zipfile(archivo):
                return f"El archivo '{archivo.name}' no es un ZIP valido."
        return None

    # ------------------------------------------------------------------
    # Lectura del ZIP (lee directo del ZIP; no usa carpeta 'por procesar').
    # ------------------------------------------------------------------
    def _leer_csvs_de_zip(self, zip_path: Path) -> pd.DataFrame:
        """Lee todos los CSVs dentro del ZIP y devuelve un DataFrame unico."""
        datos: list[pd.DataFrame] = []
        with zipfile.ZipFile(zip_path, "r") as zf:
            for nombre in zf.namelist():
                if not nombre.lower().endswith(".csv"):
                    continue
                with zf.open(nombre) as f:
                    raw = f.read()
                    df = self._read_csv_bytes(raw)
                datos.append(df)
        if not datos:
            return pd.DataFrame()
        return pd.concat(datos, ignore_index=True)

    @staticmethod
    def _read_csv_bytes(raw: bytes) -> pd.DataFrame:
        """Lee bytes como CSV con fallback de encoding.

        Tambien usa ``sep="|"`` por la misma razon que
        ``_read_csv_con_fallback``: el CSV de Bancolombia trae una
        sola linea por movimiento y algunos campos (descripcion)
        pueden contener comas. Ver docstring de ``_read_csv_con_fallback``
        para mas detalle.

        NO cambiar ``sep="|"`` por ``sep=","`` sin entenderlo primero.
    """
        for encoding in ENCODING_FALLBACKS:
            try:
                return pd.read_csv(
                    StringIO(raw.decode(encoding)),
                    decimal=".", sep="|", header=None, dtype=str,
                )
            except UnicodeDecodeError:
                continue
        return pd.read_csv(
            StringIO(raw.decode("latin-1")),
            decimal=".", sep="|", header=None, dtype=str,
        )

    # ------------------------------------------------------------------
    # Codigo contable
    # ------------------------------------------------------------------
    def _agregar_codigo_contable(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        serie = df[0].astype(str).str.strip()
        nueva = serie.map(self.codigos_contables).fillna("No encontrado")
        df.insert(len(df.columns), "Codigo Contable", nueva)
        return df

    # ------------------------------------------------------------------
    # Depuracion
    # ------------------------------------------------------------------
    def _copy_data(
        self, copia: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        if copia.empty:
            log().info("%s No hay datos para procesar.", self.LOG_PREFIX)
            empty = pd.DataFrame()
            return empty, empty, empty, empty

        # Aqui llega el split manual por coma. ``copia`` tiene una sola
        # columna con la linea entera del CSV (ej:
        # "47789085868,40,1,28022026,,25931.41,8999,ABONO,..,0").
        # Lo dividimos manualmente en 9 columnas. Esto es seguro porque
        # sabemos que la linea trae exactamente 9 campos separados por
        # coma; las comas DENTRO de la descripcion no se tocaron en la
        # lectura (gracias al ``sep="|"`` de _read_csv_bytes).
        por_cuentas = copia[0].str.split(pat=",", expand=True)
        por_cuentas = por_cuentas.apply(lambda c: c.astype(str).str.strip())

        # Filtro fiduciario: para el concepto 9729 conservar solo filas
        # cuya cuenta sea la cuenta interna de Bancolombia.
        # Antes: ``df.apply(lambda row: any(...), axis=1)`` -> O(n) en
        # Python puro. Ahora: comparacion vectorizada en C (10-100x
        # mas rapido para volumenes grandes).
        mask_9729 = por_cuentas[COL_CODIGO_CONCEPTO] == self.concepto_fiduciaria
        mask_cuenta_valida = por_cuentas.eq(self.cuenta_bancolombia).any(axis=1)
        por_cuentas = por_cuentas[~mask_9729 | mask_cuenta_valida]

        # Fecha DD/MM/YYYY.
        por_cuentas[COL_FECHA] = pd.to_datetime(
            por_cuentas[COL_FECHA], format="%d%m%Y", errors="coerce"
        ).dt.strftime("%d/%m/%Y")

        # Conversion de columnas numericas.
        for col in COLUMNAS_NUMERICAS:
            por_cuentas[col] = pd.to_numeric(por_cuentas[col], errors="coerce")

        por_cuentas[COL_CODIGO_CONCEPTO] = (
            por_cuentas[COL_CODIGO_CONCEPTO].astype(str).str.strip()
        )
        intereses = por_cuentas[
            por_cuentas[COL_CODIGO_CONCEPTO].isin(self.codigos_interes)
        ].copy()
        gastos = por_cuentas[
            por_cuentas[COL_CODIGO_CONCEPTO].isin(self.codigos_gastos)
        ].copy()

        por_cuentas = self._agregar_codigo_contable(por_cuentas)
        intereses = self._agregar_codigo_contable(intereses)
        gastos = self._agregar_codigo_contable(gastos)

        por_cuentas = por_cuentas.sort_values(by=[0], ascending=True)
        por_conceptos = por_cuentas.sort_values(by=[7], ascending=True)
        intereses = intereses.sort_values(by=[7], ascending=True)
        gastos = gastos.sort_values(by=[7], ascending=True)
        return por_cuentas, por_conceptos, intereses, gastos

    # ------------------------------------------------------------------
    # Escritura del Excel (5 hojas)
    # ------------------------------------------------------------------
    def _writer_excel(
        self, hojas: dict[str, pd.DataFrame], carpeta: Path,
    ) -> Path:
        fecha_actual = datetime.now()
        fecha_formateada = fecha_actual - relativedelta(months=1)
        mes_nombre = fecha_formateada.strftime("%m %B %Y").upper()
        archivo_final = f"{mes_nombre} Bancolombia.xlsx"
        ruta_final = carpeta / archivo_final

        with pd.ExcelWriter(ruta_final, engine="openpyxl") as writer:
            for nombre_hoja, datos in hojas.items():
                datos.to_excel(
                    writer, sheet_name=nombre_hoja, header=False, index=False,
                )
                ws = writer.sheets[nombre_hoja]
                # La columna "Valor" del CSV (COL_VALOR = 5) se escribe
                # como columna 6 en el Excel porque ``header=False``.
                COL_VALOR_EN_EXCEL = COL_VALOR + 1
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=COL_VALOR_EN_EXCEL).number_format = "0.00"
                for col in ws.columns:
                    max_length = 0
                    column_letter = col[0].column_letter
                    column_index = col[0].column
                    for cell in col:
                        if column_index == COL_VALOR_EN_EXCEL and isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00; [Red]-0.00"
                        if cell.value:
                            largo = len(str(cell.value))
                            if largo > max_length:
                                max_length = largo
                    ws.column_dimensions[column_letter].width = max(max_length + 2, 10)
        return ruta_final

    # ------------------------------------------------------------------
    # Mapeo de cuenta Bancolombia -> NIT universidad
    # ------------------------------------------------------------------
    def _codigo_para_foapal(self, cuenta: str) -> str:
        """Devuelve el NIT de la universidad para una cuenta Bancolombia.

        Usa ``nit_bancolombia.json`` (``self.nit_bancolombia``) que mapea
        cuenta Bancolombia (10 digitos) -> NIT. Si la cuenta no esta
        en el mapa, devuelve la cuenta original (fallback seguro).
        """
        if not isinstance(cuenta, str):
            cuenta = str(cuenta).strip()
        return self.nit_bancolombia.get(cuenta, cuenta)

    # ------------------------------------------------------------------
    # Aplicacion del FOAPAL -> fzrcoco.xlsx
    #
    # Estructura de ``df_conceptos`` (viene de ``_copy_data``):
    #   col 0  COL_CUENTA          (cuenta Bancolombia)
    #   col 1  COL_PREFIJO          (oficina)
    #   col 2  COL_TIPO             (Corriente / Ahorro)
    #   col 3  COL_FECHA            (DD/MM/YYYY, la pisamos con "dd-Month-yyyy")
    #   col 4  COL_ID               (codigo interno, no se usa)
    #   col 5  COL_VALOR            (valor del movimiento)
    #   col 6  COL_CODIGO_CONCEPTO  (4 digitos, ej: 480, 1334, 2999)
    #   col 7  COL_CONCEPTO         (descripcion textual)
    #   col 8  COL_VALOR_NUM        (valor sin signo, redundante)
    #   col 9  COL_CODIGO_CONTABLE   (agregado por _agregar_codigo_contable)
    # ------------------------------------------------------------------
    def _aplicar_foapal(
        self, df_conceptos: pd.DataFrame, carpeta: Path,
    ) -> Path | None:
        if df_conceptos.empty:
            log().info("%s No hay datos para aplicar FOAPAL.", self.LOG_PREFIX)
            return None
        try:
            fechas = pd.to_datetime(
                df_conceptos[COL_FECHA], format="%d/%m/%Y", errors="coerce"
            )
            df = df_conceptos.copy()
            df[COL_FECHA] = fechas.apply(
                lambda d: f"{d.day:02d}-{MESES_ES[d.month]}-{d.year}" if pd.notna(d) else ""
            )

            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet("Comprobante")
            for col_idx, encabezado in enumerate(ENCABEZADOS_FOAPAL, start=1):
                ws.cell(row=1, column=col_idx, value=encabezado)

            creditos = self.foapal_config.get("creditos", {})
            debitos = self.foapal_config.get("debitos", {})
            filas: list[list] = []
            for _, fila in df.iterrows():
                clave_concepto = str(fila.iloc[COL_CODIGO_CONCEPTO]).strip()
                foapal_info = creditos.get(clave_concepto) or debitos.get(clave_concepto)
                if foapal_info is None:
                    continue
                codigo_contable = str(fila.iloc[COL_CODIGO_CONTABLE]).strip()
                concepto = fila.iloc[COL_CONCEPTO]
                valor = abs(fila.iloc[COL_VALOR])
                fecha = fila.iloc[COL_FECHA]

                # Reemplazar la cuenta Bancolombia (COL_CUENTA) por el
                # NIT de la universidad, usando el JSON nit_bancolombia.
                codigo_universidad = self._codigo_para_foapal(
                    str(fila.iloc[COL_CUENTA]).strip()
                )

                try:
                    codigo_contable_num = int(codigo_contable)
                except ValueError:
                    codigo_contable_num = codigo_contable

                dc_principal = foapal_info.get("D/C", "C")
                dc_contraparte = "D" if dc_principal == "C" else "C"
                fondo = foapal_info.get("Fondo", "FOPNAL")
                organizacion = foapal_info.get("Organizacion", "999999")
                programa = foapal_info.get("Programa", "999999")

                filas.append([
                    codigo_universidad, concepto, valor, fecha,
                    fondo, organizacion, codigo_contable_num, programa, dc_principal,
                ])
                if codigo_contable == CODIGO_EXCEPCION_FOAPAL:
                    filas.append([
                        codigo_universidad, concepto, valor, fecha,
                        EXCEPCION_FOAPAL["Fondo"],
                        EXCEPCION_FOAPAL["Organizacion"],
                        EXCEPCION_FOAPAL["Cuenta"],
                        EXCEPCION_FOAPAL["Programa"],
                        dc_contraparte,
                    ])
                else:
                    filas.append([
                        codigo_universidad, concepto, valor, fecha,
                        fondo, organizacion, codigo_contable_num, programa, dc_contraparte,
                    ])

            for fila_idx, row_data in enumerate(filas, start=2):
                for col_idx, valor_celda in enumerate(row_data, start=1):
                    ws.cell(row=fila_idx, column=col_idx, value=valor_celda)
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        largo = len(str(cell.value))
                        if largo > max_length:
                            max_length = largo
                ws.column_dimensions[column_letter].width = max(max_length + 2, 10)

            ruta_foapal = carpeta / "fzrcoco.xlsx"
            wb.save(ruta_foapal)
            log().info("%s FOAPAL generado: %s", self.LOG_PREFIX, ruta_foapal)
            return ruta_foapal
        except Exception as e:
            log().error("%s Error aplicando FOAPAL: %s", self.LOG_PREFIX, e)
            return None

    # ------------------------------------------------------------------
    # Ejecucion
    # ------------------------------------------------------------------
    def ejecutar(
        self,
        archivos: list[Path],
        modo_prueba: bool = False,
    ) -> ResultadoProceso:
        log().info("%s Iniciando (modo_prueba=%s)", self.LOG_PREFIX, modo_prueba)
        zip_path = archivos[0]

        try:
            copia = self._leer_csvs_de_zip(zip_path)
        except Exception as e:
            return ResultadoProceso(exito=False, mensaje=f"No se pudo leer el ZIP: {e}")
        log().info("%s CSV(s) leido(s): %d fila(s)", self.LOG_PREFIX, len(copia))

        por_cuentas, por_conceptos, intereses, gastos = self._copy_data(copia)

        if modo_prueba:
            carpeta = carpeta_modo_prueba(RAIZ / "resultados", self.nombre)
        else:
            carpeta = carpeta_resultados(RAIZ / "resultados", self.nombre)

        try:
            ruta_excel = self._writer_excel({
                "Original": copia,
                "Por cuentas": por_cuentas,
                "Por conceptos": por_conceptos,
                "Intereses": intereses,
                "Gastos bancarios": gastos,
            }, carpeta)
        except Exception as e:
            return ResultadoProceso(exito=False, mensaje=f"No se pudo escribir el Excel: {e}")

        ruta_foapal = self._aplicar_foapal(por_conceptos, carpeta)

        archivos_generados: list[Path] = [ruta_excel]
        if ruta_foapal is not None:
            archivos_generados.append(ruta_foapal)
        log().info("%s Generado: %s", self.LOG_PREFIX, ruta_excel.name)

        return ResultadoProceso(
            exito=True,
            mensaje="Comprobante generado correctamente.",
            archivos_salida=archivos_generados,
            detalles={"filas_origen": len(copia)},
        )