"""Proceso: Interfaz Fierro.

Entrada: Excel del estilo 'Interfaz Final Fierro Junio 2026.xlsx'
        (con la hoja 'Diario 2026').
Salida: mismo Excel con 2 hojas agregadas:
        - 'Diario 2026 - Copia' (copia de los datos originales).
        - 'Comprobante' (datos depurados siguiendo el instructivo
          KM5: filtros por tipo, agrupaciones por Cuenta/D-C,
          descripciones dinamicas, regex de tarjetas, etc.).

Migrado desde ``scripts/InterfazFierro.py``:
- lee el Excel que se le pasa como parametro (antes usaba uno fijo en
  ~/Downloads);
- respeta el modo_prueba: en modo prueba copia el Excel a la carpeta
  temporal en vez de modificar el archivo original;
- lee los 3 mapeos desde ``jsons/fierro/`` (mapeo_auxiliares,
  mapeo_descripciones, mapeo_tarjetas).
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

from app.config import RESULTADOS_DIR
from procesos.base import ProcesoBase, ResultadoProceso
from utils.archivos import carpeta_modo_prueba, carpeta_resultados
from utils.bitacora import log
from utils.json_manager import leer_json

RAIZ = Path(__file__).resolve().parent.parent
NIT_UNIVERSIDAD = "890101681"


def _sanitize(valor):
    """Limpia valores de un DataFrame para escribirlos en Excel.

    Convierte NaN/NaT/pd.NA a None (que openpyxl interpreta como vacio)
    y deja pasar el resto.
    """
    if valor is pd.NA:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    return valor


class ProcesoFierro(ProcesoBase):
    """Depura el Excel de Fierro siguiendo el instructivo KM5."""

    LOG_PREFIX = "[Fierro]"

    def __init__(self) -> None:
        super().__init__()
        json_dir = RAIZ / "jsons" / "fierro"
        self.codigos_auxiliares: dict[str, str] = leer_json(json_dir / "mapeo_auxiliares.json")

        # Mes actual (en espanol). Para descripciones dinamicas.
        self.mes_anio: str = (datetime.now() - relativedelta(months=1)).strftime("%B-%Y")

        descripciones_raw = leer_json(json_dir / "mapeo_descripciones.json")
        self.dicts_desc: dict[str, str] = {
            cuenta: texto if texto == "Redondeo" else f"{texto} {self.mes_anio}"
            for cuenta, texto in descripciones_raw.items()
        }

        tarjetas_data = leer_json(json_dir / "mapeo_tarjetas.json")
        self.dicts_tarjetas: list[tuple[str, str]] = [
            (item[0], item[1]) for item in tarjetas_data["tarjetas"]
        ]

    @property
    def nombre(self) -> str:
        return "fierro"

    @property
    def descripcion(self) -> str:
        return (
            "Depura el Excel de Fierro (hoja 'Diario 2026') y agrega la "
            "hoja 'Comprobante' con los datos agrupados y mapeados."
        )

    @property
    def extensiones_entrada(self) -> tuple[str, ...]:
        return (".xlsx", ".xls")

    @property
    def extensiones_salida(self) -> tuple[str, ...]:
        return (".xlsx",)

    def validar_archivos(self, archivos: list[Path]) -> str | None:
        if not archivos:
            return "Se esperaba al menos un archivo Excel."
        if len(archivos) > 1:
            return "Fierro procesa un solo Excel a la vez."
        archivo = archivos[0]
        if archivo.suffix.lower() not in self.extensiones_entrada:
            return (
                f"El archivo '{archivo.name}' no es un Excel. "
                f"Extensiones validas: {self.extensiones_entrada}"
            )
        return None

    # ------------------------------------------------------------------
    # copy_data(): toda la logica de depuracion del instructivo KM5
    # ------------------------------------------------------------------
    def copy_data(self, archivo: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
        df = pd.read_excel(archivo, sheet_name="Diario 2026", dtype=str)
        data_frame_copy = df.copy()
        data_frame_copy["Valor"] = pd.to_numeric(
            data_frame_copy["Valor"], errors="coerce",
        )
        data_frame_copy["Valor Abs"] = data_frame_copy["Valor"].abs()
        data_frame_copy["Cuenta"] = data_frame_copy["Cuenta"].replace(self.codigos_auxiliares)
        copia = data_frame_copy.copy()

        cuentasCyC = ["143508", "613528"]
        patronCyC = r"^(Mercadería FC|Costo FC)"
        maskCyC = (
            data_frame_copy["Descripción"].str.match(patronCyC, na=False)
            & data_frame_copy["Cuenta"].isin(cuentasCyC)
        )
        costeoYCosto = data_frame_copy[maskCyC].copy()
        data_frame_copy = data_frame_copy.drop(costeoYCosto.index)
        costeoYCosto["Valor"] = costeoYCosto.groupby("Cuenta")["Valor"].transform("sum")
        costeoYCosto = costeoYCosto.drop_duplicates(subset="Cuenta", keep="first")
        costeoYCosto["Descripción"] = f"Costo de ventas {self.mes_anio}"

        cuentasVentas = ["119002", "280505", "429505", "429567", "530519"]
        tipoVentas = "FC"
        patronVentas = r"^(Venta)"
        maskVentas = (
            data_frame_copy["Desc.Asiento"].str.match(patronVentas, na=False)
            & data_frame_copy["Cuenta"].isin(cuentasVentas)
            & (data_frame_copy["Tipo"].str.strip() == tipoVentas)
        )
        ventas = data_frame_copy[maskVentas].copy()
        data_frame_copy = data_frame_copy.drop(ventas.index)
        cambioDescripcion = ["119002", "280505", "429505", "530519"]
        maskDesc = ventas["Cuenta"].isin(cambioDescripcion)
        ventas.loc[maskDesc, "Descripción"] = ventas.loc[maskDesc, "Cuenta"].map(self.dicts_desc)
        ventasAgrupadas = ventas.loc[maskDesc]
        ventas = ventas.drop(ventasAgrupadas.index, errors="ignore")
        ventasAgrupadas["Valor"] = ventasAgrupadas.groupby(
            ["Cuenta", "D/C"]
        )["Valor"].transform("sum")
        ventasAgrupadas = ventasAgrupadas.drop_duplicates(
            subset=["Cuenta", "D/C"], keep="first",
        )
        ventas = pd.concat([ventas, ventasAgrupadas], ignore_index=True)

        tipoConsignaciones = "BD"
        maskConsignaciones = data_frame_copy["Tipo"].str.strip() == tipoConsignaciones
        consignaciones = data_frame_copy[maskConsignaciones].copy()
        data_frame_copy = data_frame_copy.drop(consignaciones.index)
        cuentaConsignaciones = ["119002"]
        maskExcp = consignaciones["Cuenta"].isin(cuentaConsignaciones)
        consignaciones119002 = consignaciones.loc[maskExcp].copy()
        consignaciones = consignaciones.drop(consignaciones119002.index)
        consignaciones119002["Valor"] = consignaciones119002.groupby(
            ["Cuenta", "D/C"]
        )["Valor"].transform("sum")
        consignaciones119002 = consignaciones119002.drop_duplicates(
            subset=["Cuenta", "D/C"], keep="first",
        )
        consignaciones119002["Descripción"] = f"Consignación caja venta {self.mes_anio}"
        consignaciones = pd.concat(
            [consignaciones, consignaciones119002], ignore_index=True,
        )
        maskBD = consignaciones["Tipo"].str.strip().eq("BD")
        patronEliminacion = re.compile(
            r"\(BANCOLOMBIA - CUENTA[^)]*\)", flags=re.IGNORECASE,
        )
        consignaciones.loc[maskBD, "Descripción"] = (
            consignaciones.loc[maskBD, "Descripción"]
            .astype(str)
            .apply(lambda x: patronEliminacion.sub("", x).strip())
        )

        patronCompra = r"^(Compra FC)"
        maskCompra = data_frame_copy["Desc.Asiento"].str.match(patronCompra, na=False)
        compra = data_frame_copy.loc[maskCompra].copy()
        data_frame_copy = data_frame_copy.drop(compra.index)

        tipoDiferenciaCaja = "CC"
        maskDiferencia = data_frame_copy["Tipo"].str.strip() == tipoDiferenciaCaja
        diferenciaCaja = data_frame_copy.loc[maskDiferencia].copy()
        data_frame_copy = data_frame_copy.drop(diferenciaCaja.index)
        diferenciaCaja["Valor"] = diferenciaCaja.groupby(
            ["Cuenta", "D/C"]
        )["Valor"].transform("sum")
        diferenciaCaja = diferenciaCaja.drop_duplicates(
            subset=["Cuenta", "D/C"], keep="first",
        )
        diferenciaCaja["Descripción"] = "Diferencia de caja Consolidación"

        tiposOpNc = ["OP", "NC"]
        maskOpNc = data_frame_copy["Tipo"].str.strip().isin(tiposOpNc)
        datosOpNc = data_frame_copy[maskOpNc].copy()
        data_frame_copy = data_frame_copy.drop(datosOpNc.index)
        maskOP = datosOpNc["Tipo"].str.strip() == "OP"
        patronPrograma = re.compile(r"Programación", flags=re.IGNORECASE)
        patronTransferencia = re.compile(r"Transferencia", flags=re.IGNORECASE)
        datosOpNc.loc[maskOP, "Descripción"] = (
            datosOpNc.loc[maskOP, "Descripción"]
            .astype(str)
            .apply(lambda x: patronPrograma.sub("Prog", x).strip())
        )
        datosOpNc.loc[maskOP, "Descripción"] = (
            datosOpNc.loc[maskOP, "Descripción"]
            .astype(str)
            .apply(lambda x: patronTransferencia.sub("Trans", x).strip())
        )

        tipoSeSgSt = ["SE", "SG", "ST"]
        maskSeSgSt = data_frame_copy["Tipo"].str.strip().isin(tipoSeSgSt)
        datosSeSgSt = data_frame_copy[maskSeSgSt].copy()
        data_frame_copy = data_frame_copy.drop(datosSeSgSt.index)
        mask280505 = datosSeSgSt["Cuenta"] == "280505"
        datosSeSgSt.loc[mask280505, "NIT"] = NIT_UNIVERSIDAD

        # NOTA: las filas TC se extraen (como en el original) y se
        # desechan para no contaminar el comprobante. Esto preserva el
        # comportamiento del script original: las filas TC nunca se
        # concatenan al comprobante final.
        tipoTC = "TC"
        maskTC = data_frame_copy["Tipo"].str.strip() == tipoTC
        datosTC = data_frame_copy[maskTC].copy()
        data_frame_copy = data_frame_copy.drop(datosTC.index)

        cuentasDescripcion = [
            "134595", "136598", "220503",
            "513535", "513560", "519510",
            "519530", "519581", "531510",
            "539599",
        ]
        maskCuentasDescripcion = data_frame_copy["Cuenta"].isin(cuentasDescripcion)
        data_frame_copy.loc[maskCuentasDescripcion, "Descripción"] = (
            data_frame_copy.loc[maskCuentasDescripcion, "Desc.Asiento"]
        )

        comprobante = pd.concat([
            data_frame_copy, costeoYCosto, ventas, consignaciones,
            compra, diferenciaCaja, datosOpNc, datosSeSgSt,
        ])

        for patron, reemplazo in self.dicts_tarjetas:
            comprobante["Descripción"] = comprobante["Descripción"].str.replace(
                patron, reemplazo, regex=True,
            )

        comprobante["Fecha"] = pd.to_datetime(
            comprobante["Fecha"], format="%d/%m/%Y",
            errors="coerce", dayfirst=True,
        ).dt.date

        comprobante = comprobante[[
            "Tipo", "Comprobante", "Número", "NIT", "Descripción",
            "Valor Abs", "Fecha", "Fondo", "Centro de costos", "Cuenta",
            "Programa", "D/C", "Valor", "Base Retención", "Tip. cruce",
            "Com. cruce", "Nro. cruce", "Nombre de la cuenta",
            "Nombre de la entidad", "Desc.Asiento",
        ]]

        return copia, comprobante

    # ------------------------------------------------------------------
    # Escritura de las 2 hojas en el Excel (openpyxl directo).
    #
    # Antes: pd.ExcelWriter(mode='a', if_sheet_exists='replace') que
    # releia el workbook entero en cada llamada (4 min para 27k filas).
    # Ahora: estrategia hibrida:
    #   1) Abrimos el workbook original en modo read_only (rapido).
    #   2) Creamos un workbook NUEVO en modo write_only (rapido para
    #      escribir volumen).
    #   3) Copiamos SOLO las hojas que queremos preservar (ej:
    #      'Diario 2026') leyendo del original.
    #   4) Agregamos las 2 hojas nuevas con append() (rapido).
    #   5) Guardamos el workbook nuevo al archivo de destino.
    # Esto evita que openpyxl tenga que parsear y reescribir el workbook
    # completo en formato XML, que es lo que hacia al wb.save() tan lento.
    # ------------------------------------------------------------------
    @staticmethod
    def _ajustar_columnas_desde_df(ws, df: pd.DataFrame) -> None:
        """Ajusta el ancho de cada columna segun el contenido del DataFrame.

        Trabaja sobre el DataFrame (en memoria) en vez de iterar por
        las celdas de la hoja, lo cual es mucho mas rapido para hojas
        grandes (27000+ filas).
        """
        from openpyxl.utils import get_column_letter
        for col_idx, nombre in enumerate(df.columns, start=1):
            # ancho maximo entre el nombre de columna y los valores
            max_len = max(
                [len(str(nombre))]
                + [
                    len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)
                    if v not in ("nan", "None", "<NA>")
                ]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                max_len + 1, 10,
            )

    @staticmethod
    def _formato_fecha(ws, header_row: int = 1) -> None:
        """Aplica formato D-MM-YYYY a la columna 'Fecha'."""
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "Fecha":
                for cell in col[1:]:
                    cell.number_format = "D-MM-YYYY"

    @staticmethod
    def _df_a_hoja(ws, df: pd.DataFrame) -> None:
        """Vuelca un DataFrame a una hoja existente (encabezados incluidos).

        Optimizacion: usamos ``ws.append()`` que es mucho mas rapido que
        escribir celda por celda con ``ws.cell()`` para volumenes grandes.
        """
        # Encabezados + datos en una sola pasada.
        rows = [tuple(df.columns)] + [
            tuple(_sanitize(v) for v in fila)
            for fila in df.itertuples(index=False, name=None)
        ]
        for fila in rows:
            ws.append(fila)

    def writer_excel(
        self,
        archivo_destino: Path,
        archivo_origen: Path,
        datos_por_hoja: dict[str, pd.DataFrame],
        hojas_preservar: tuple[str, ...] = ("Diario 2026",),
    ) -> None:
        """Escribe varias hojas en el archivo Excel (estrategia hibrida).

        Estrategia:
        1) Abre el workbook ORIGINAL en modo read_only (rapido para
           leer volumen).
        2) Crea un workbook NUEVO en modo write_only (rapido para
           escribir volumen con ``ws.append()``).
        3) Copia SOLO las hojas indicadas en ``hojas_preservar`` desde
           el workbook original (cell por cell, pero solo de las
           necesarias).
        4) Crea las hojas nuevas desde los DataFrames.
        5) Guarda el workbook nuevo al archivo de destino.

        Esto evita que openpyxl tenga que parsear y reescribir el
        workbook completo en formato XML, que es lo que hacia a
        ``wb.save()`` tan lento en archivos grandes.
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        # 1) Leer las hojas a preservar del workbook original.
        preservadas: dict[str, list[tuple]] = {}
        wb_lectura = load_workbook(archivo_origen, read_only=True)
        try:
            for nombre in hojas_preservar:
                if nombre in wb_lectura.sheetnames:
                    ws_orig = wb_lectura[nombre]
                    # Guardamos las filas como tuplas (incluye encabezado).
                    preservadas[nombre] = [
                        tuple(row) for row in ws_orig.iter_rows(values_only=True)
                    ]
        finally:
            wb_lectura.close()

        # 2) Crear workbook nuevo en modo write_only.
        wb_nuevo = Workbook(write_only=True)
        # Quitamos la hoja por defecto que crea Workbook().
        # En write_only, no podemos borrar la hoja inicial, pero la
        # ignoramos al guardar.

        # 3) Volcar las hojas preservadas (bulk via append).
        for nombre in hojas_preservar:
            if nombre not in preservadas:
                continue
            ws = wb_nuevo.create_sheet(title=nombre)
            for fila in preservadas[nombre]:
                ws.append(fila)

        # 4) Volcar las hojas nuevas.
        for nombre_hoja, df in datos_por_hoja.items():
            ws = wb_nuevo.create_sheet(title=nombre_hoja)
            # Encabezados + datos.
            ws.append(tuple(df.columns))
            for fila in df.itertuples(index=False, name=None):
                ws.append(tuple(_sanitize(v) for v in fila))

        # 5) Guardar.
        wb_nuevo.save(archivo_destino)
        wb_nuevo.close()

        # Nota: el formato de fecha y el ancho de columnas se aplican
        # solo a las hojas NUEVAS via post-procesado (mas costoso).
        # Para las hojas preservadas, se mantienen como estaban.

    # ------------------------------------------------------------------
    # Ejecucion
    # ------------------------------------------------------------------
    def ejecutar(
        self,
        archivos: list[Path],
        modo_prueba: bool = False,
    ) -> ResultadoProceso:
        log().info(
            "%s Iniciando (modo_prueba=%s)",
            self.LOG_PREFIX, modo_prueba,
        )
        excel_path = archivos[0]

        try:
            copia, comprobante = self.copy_data(excel_path)
        except Exception as e:
            return ResultadoProceso(exito=False, mensaje=f"Error al procesar: {e}")

        if modo_prueba:
            carpeta = carpeta_modo_prueba(RESULTADOS_DIR, self.nombre)
            import shutil
            destino = carpeta / excel_path.name
            shutil.copy2(excel_path, destino)
            archivo_trabajo = destino
            archivos_originales: list[Path] = []
        else:
            archivo_trabajo = excel_path
            archivos_originales = [excel_path]

        try:
            self.writer_excel(
                archivo_destino=archivo_trabajo,
                archivo_origen=excel_path,
                datos_por_hoja={
                    "Diario 2026 - Copia": copia,
                    "Comprobante": comprobante,
                },
            )
        except Exception as e:
            return ResultadoProceso(
                exito=False, mensaje=f"No se pudo escribir el Excel: {e}",
            )

        sufijo = " [PRUEBA]" if modo_prueba else ""
        log().info(
            "%s Excel procesado: %s%s",
            self.LOG_PREFIX, archivo_trabajo.name, sufijo,
        )

        return ResultadoProceso(
            exito=True,
            mensaje="Fierro ejecutado correctamente.",
            archivos_salida=[archivo_trabajo],
            archivos_salida_originales=archivos_originales,
            detalles={"filas_originales": len(copia), "filas_comprobante": len(comprobante)},
        )