"""Proceso: Interfaz Zeus.

Entrada: Excel 'Interfaz Final Zeus.xlsx' (con la hoja 'Exportar').
Salida: mismo Excel con 2 hojas agregadas:
        - 'Exportar - Copia' (datos sin modificar, con 'Cuenta1' depurada).
        - 'Depurado' (datos con Valor2, BaseAbs, Tarifa y las
          agrupaciones por Nit/Cuenta1/Fecha aplicadas).

Migrado desde ``scripts/InterfazZeus.py``:
- lee el Excel que se le pasa como parametro (antes usaba uno fijo en
  ~/Downloads);
- respeta el modo_prueba: en modo prueba copia el Excel a la carpeta
  temporal en vez de modificar el original;
- lee los auxiliares desde ``jsons/zeus/auxiliares_zeus.json``.
"""
from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd

from procesos.base import ProcesoBase, ResultadoProceso
from utils.archivos import carpeta_modo_prueba, carpeta_resultados
from utils.bitacora import log
from utils.json_manager import leer_json

RAIZ = Path(__file__).resolve().parent.parent


class ProcesoZeus(ProcesoBase):
    """Depura el Excel de Zeus: cuentas 8 digitos -> 6 digitos,
    calcula Valor2/BaseAbs/Tarifa y agrupa por Nit/Cuenta1/Fecha."""

    LOG_PREFIX = "[Zeus]"

    def __init__(self) -> None:
        super().__init__()
        json_dir = RAIZ / "jsons" / "zeus"
        aux_data = leer_json(json_dir / "auxiliares_zeus.json")
        self.aux: dict[str, str] = {
            patron: remplazo for patron, remplazo in aux_data["auxiliares"]
        }

    @property
    def nombre(self) -> str:
        return "zeus"

    @property
    def descripcion(self) -> str:
        return (
            "Depura el Excel de Zeus (hoja 'Exportar'): auxiliares de 8 "
            "digitos a 6 digitos, Valor2, BaseAbs, Tarifa y agrupacion "
            "por Nit/Cuenta1/Fecha. Agrega las hojas 'Exportar - Copia' "
            "y 'Depurado'."
        )

    @property
    def extensiones_entrada(self) -> tuple[str, ...]:
        return (".xlsx", ".xls")

    @property
    def extensiones_salida(self) -> tuple[str, ...]:
        return (".xlsx",)

    def validar_archivos(self, archivos: list[Path]) -> str | None:
        if not archivos:
            return "Se esperaba al menos un archivo Excel."
        if len(archivos) > 1:
            return "Zeus procesa un solo Excel a la vez."
        archivo = archivos[0]
        if archivo.suffix.lower() not in self.extensiones_entrada:
            return (
                f"El archivo '{archivo.name}' no es un Excel. "
                f"Extensiones validas: {self.extensiones_entrada}"
            )
        return None

    # ------------------------------------------------------------------
    # Depuracion: migracion literal del script original
    # ------------------------------------------------------------------
    def copy_data(self, archivo: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
        # Detectar la hoja que tiene la columna 'Cuenta1'.
        # El script original tomaba siempre la primera hoja, lo que falla
        # en algunos Excels donde la primera hoja es auxiliar.
        xls = pd.ExcelFile(archivo)
        hoja_objetivo = None
        for hoja in xls.sheet_names:
            head = pd.read_excel(xls, sheet_name=hoja, dtype=str, nrows=0)
            if "Cuenta1" in head.columns:
                hoja_objetivo = hoja
                break
        if hoja_objetivo is None:
            raise ValueError(
                f"No se encontro la columna 'Cuenta1' en ninguna hoja "
                f"del archivo {archivo.name}."
            )
        log().info("%s Hoja objetivo: %s", self.LOG_PREFIX, hoja_objetivo)
        data_frame_copia = pd.read_excel(archivo, sheet_name=hoja_objetivo, dtype=str)

        # Pasar auxiliares (Cuenta1) de 8 digitos a 6 digitos.
        for patron, remplazo in self.aux.items():
            data_frame_copia["Cuenta1"] = data_frame_copia["Cuenta1"].str.replace(
                patron, remplazo, regex=True,
            )

        copia = data_frame_copia.copy()

        # Columnas numericas y fecha.
        numericos = ["valor", "Base"]
        for col in numericos:
            copia[col] = pd.to_numeric(copia[col], errors="coerce")
        copia["Fecha"] = pd.to_datetime(
            copia["Fecha"], errors="coerce", dayfirst=False,
        ).dt.date

        # Columna Valor 2: negativo si Tipo_Movto == 'C'.
        maskC = copia["Tipo_Movto"].str.strip() == "C"
        copia["Valor2"] = np.where(maskC, -copia["valor"], copia["valor"])
        copia["BaseAbs"] = copia["Base"].abs()

        # Base de retenciones.
        cuentasRetenciones = re = __import__("re").compile(r"(?:2365|2367|2368)")
        maskRetenciones = copia["Cuenta1"].str.match(cuentasRetenciones)
        copia["Tarifa"] = np.where(
            maskRetenciones & (copia["BaseAbs"] != 0),
            (copia["valor"] / copia["BaseAbs"]) * 100,
            np.nan,
        ).round(2)
        copia.loc[maskRetenciones, "Concepto"] = copia.loc[
            maskRetenciones, "BaseAbs"
        ].astype(str)

        # Agrupacion por Nit / Cuenta1 / Fecha.
        copia["Cuenta1"] = copia["Cuenta1"].str.strip()
        cuentasPorAgrupar = [
            "7101", "7105", "119006", "134597", "134598", "143507",
            "280505", "280523", "530519", "613528",
        ]
        maskCuentasAgrupadas = copia["Cuenta1"].isin(cuentasPorAgrupar)
        copia.loc[maskCuentasAgrupadas, "Valor"] = copia.groupby(
            ["Nit", "Cuenta1", "Fecha"],
        )[["valor", "Valor2"]].transform("sum")
        copia.loc[maskCuentasAgrupadas] = copia.drop_duplicates(
            subset=["Nit", "Cuenta1", "Fecha"], keep="first",
        )

        return data_frame_copia, copia

    # ------------------------------------------------------------------
    # Escritura de las 2 hojas en el Excel (estrategia hibrida write_only).
    #
    # Misma estrategia que en ``procesos/fierro.py``: en lugar de
    # ``pd.ExcelWriter(mode='a')`` (que relee el workbook entero en
    # cada llamada y es muy lento para archivos grandes), abrimos el
    # workbook original en modo ``read_only``, creamos uno NUEVO en
    # modo ``write_only``, copiamos solo las hojas que queremos
    # preservar (``Exportar``) y agregamos las 2 hojas nuevas.
    # ------------------------------------------------------------------
    def writer_excel(
        self,
        archivo_destino: Path,
        archivo_origen: Path,
        datos_por_hoja: dict[str, pd.DataFrame],
        hojas_preservar: tuple[str, ...] = ("Exportar",),
    ) -> None:
        from openpyxl import Workbook

        # 1) Leer las hojas a preservar del workbook original.
        preservadas: dict[str, list[tuple]] = {}
        from openpyxl import load_workbook
        wb_lectura = load_workbook(archivo_origen, read_only=True)
        try:
            for nombre in hojas_preservar:
                if nombre in wb_lectura.sheetnames:
                    ws_orig = wb_lectura[nombre]
                    preservadas[nombre] = [
                        tuple(row)
                        for row in ws_orig.iter_rows(values_only=True)
                    ]
        finally:
            wb_lectura.close()

        # 2) Crear workbook nuevo en modo write_only.
        wb_nuevo = Workbook(write_only=True)

        # 3) Volcar las hojas preservadas.
        for nombre in hojas_preservar:
            if nombre not in preservadas:
                continue
            ws = wb_nuevo.create_sheet(title=nombre)
            for fila in preservadas[nombre]:
                ws.append(fila)

        # 4) Volcar las hojas nuevas (Exportar - Copia, Depurado).
        def _sanitize(v):
            if v is pd.NA:
                return None
            try:
                if pd.isna(v):
                    return None
            except (TypeError, ValueError):
                pass
            return v

        for nombre_hoja, df in datos_por_hoja.items():
            ws = wb_nuevo.create_sheet(title=nombre_hoja)
            ws.append(tuple(df.columns))
            for fila in df.itertuples(index=False, name=None):
                ws.append(tuple(_sanitize(v) for v in fila))

        # 5) Guardar.
        wb_nuevo.save(archivo_destino)
        wb_nuevo.close()

    # ------------------------------------------------------------------
    # Ejecucion
    # ------------------------------------------------------------------
    def ejecutar(
        self,
        archivos: list[Path],
        modo_prueba: bool = False,
    ) -> ResultadoProceso:
        log().info(
            "%s Iniciando (modo_prueba=%s)",
            self.LOG_PREFIX, modo_prueba,
        )
        excel_path = archivos[0]

        try:
            copia, depurado = self.copy_data(excel_path)
        except Exception as e:
            return ResultadoProceso(exito=False, mensaje=f"Error al procesar: {e}")

        if modo_prueba:
            carpeta = carpeta_modo_prueba(RAIZ / "resultados", self.nombre)
            import shutil
            destino = carpeta / excel_path.name
            shutil.copy2(excel_path, destino)
            archivo_trabajo = destino
            archivos_originales: list[Path] = []
        else:
            archivo_trabajo = excel_path
            archivos_originales = [excel_path]

        try:
            self.writer_excel(
                archivo_destino=archivo_trabajo,
                archivo_origen=excel_path,
                datos_por_hoja={
                    "Exportar - Copia": copia,
                    "Depurado": depurado,
                },
            )
        except Exception as e:
            return ResultadoProceso(
                exito=False, mensaje=f"No se pudo escribir el Excel: {e}",
            )

        sufijo = " [PRUEBA]" if modo_prueba else ""
        log().info(
            "%s Excel procesado: %s%s",
            self.LOG_PREFIX, archivo_trabajo.name, sufijo,
        )

        return ResultadoProceso(
            exito=True,
            mensaje="Zeus ejecutado correctamente.",
            archivos_salida=[archivo_trabajo],
            archivos_salida_originales=archivos_originales,
            detalles={"filas_originales": len(copia), "filas_depurado": len(depurado)},
        )