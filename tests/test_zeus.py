"""Test end-to-end del proceso Zeus."""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

RAIZ = Path(__file__).resolve().parent.parent
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from core.zeus import ProcesoZeus

COLS_ZEUS = [
    "Nit", "Cuenta1", "Tipo_Movto", "valor", "Base", "Fecha", "Concepto",
]


def _crear_excel_sintetico(
    ruta: Path,
    filas: list[list] | None = None,
    nombre_hoja: str = "Exportar",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    ws.append(COLS_ZEUS)
    if filas is None:
        filas = [
            ["890101681", "11902101", "D", "1500.00", "1500.00",
             "2026-01-15", "Concepto A"],
            ["890101681", "710101", "D", "300.00", "300.00",
             "2026-01-15", "Concepto B"],
        ]
    for fila in filas:
        ws.append(fila)
    wb.save(ruta)


def test_validar_archivos_excel_valido(tmp_path: Path) -> None:
    excel = tmp_path / "zeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    if proceso.EN_DESARROLLO:
        error = proceso.validar_archivos([excel])
        assert error is not None
        assert "desarrollo" in error.lower()
    else:
        assert proceso.validar_archivos([excel]) is None


def test_validar_archivos_sin_archivos() -> None:
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([])
    assert error is not None


def test_validar_archivos_demasiados(tmp_path: Path) -> None:
    a1 = tmp_path / "uno.xlsx"
    a2 = tmp_path / "dos.xlsx"
    _crear_excel_sintetico(a1)
    _crear_excel_sintetico(a2)
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([a1, a2])
    assert error is not None


def test_validar_archivos_extension_invalida(tmp_path: Path) -> None:
    txt = tmp_path / "falso.txt"
    txt.write_text("datos")
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([txt])
    assert error is not None


def test_ejecutar_modo_prueba_genera_archivo_en_carpeta_prueba(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus está en desarrollo; ejecutar() bloquea.")

    monkeypatch.setattr("core.zeus.RESULTADOS_DIR", tmp_path)

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=True)

    assert resultado.exito, resultado.mensaje
    assert len(resultado.archivos_salida) == 1
    salida = resultado.archivos_salida[0]
    assert salida.exists()
    assert "zeus" in salida.parts
    assert any(part.startswith("_prueba_") for part in salida.parts)


def test_ejecutar_modo_produccion_modifica_original(tmp_path: Path) -> None:
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus está en desarrollo; ejecutar() bloquea.")
    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito
    assert resultado.archivos_salida[0] == excel
    assert resultado.archivos_salida_originales == [excel]


def test_excel_resultado_tiene_hojas_preservadas_y_nuevas(tmp_path: Path) -> None:
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus está en desarrollo; ejecutar() bloquea.")

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        nombres = set(wb.sheetnames)
    finally:
        wb.close()
    assert "Exportar" in nombres
    assert "Exportar - Copia" in nombres
    assert "Comprobante" in nombres


def test_depurado_aplica_auxiliares_8_a_6_digitos(tmp_path: Path) -> None:
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus está en desarrollo; ejecutar() bloquea.")

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        ws = wb["Exportar - Copia"]
        filas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header = filas[0]
    idx_cuenta1 = header.index("Cuenta1")
    cuentas = [fila[idx_cuenta1] for fila in filas[1:]]
    assert "119021" in cuentas


def test_comprobante_agrega_columnas_nuevas(tmp_path: Path) -> None:
    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        ws = wb["Comprobante"]
        header = next(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    assert "Valor2" in header
    assert "BaseAbs" in header
    assert "Tarifa" in header


def test_resultado_incluye_detalles_de_filas(tmp_path: Path) -> None:
    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito
    assert resultado.detalles is not None
    assert resultado.detalles["filas_originales"] == 2
    assert resultado.detalles["filas_comprobante"] == 2


def test_ejecutar_excel_sin_hoja_exportar_retorna_error(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SinCuenta1"
    ws.append(["a", "b"])
    wb.save(tmp_path / "malo.xlsx")

    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([tmp_path / "malo.xlsx"], modo_prueba=False)
    assert not resultado.exito
    assert "exportar" in resultado.mensaje.lower()


def test_agrupacion_no_genera_filas_vacias(tmp_path: Path) -> None:
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus está en desarrollo; ejecutar() bloquea.")

    filas = [
        ["890101681", "7101", "D", "100.00", "100.00",
         "2026-01-15", "Concepto A"],
        ["890101681", "7101", "D", "200.00", "200.00",
         "2026-01-15", "Concepto B"],
        ["890101681", "11902101", "D", "1500.00", "1500.00",
         "2026-01-15", "Concepto C"],
    ]
    excel = tmp_path / "ZeusAgrupado.xlsx"
    _crear_excel_sintetico(excel, filas=filas)

    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito, resultado.mensaje

    wb = load_workbook(excel, read_only=True)
    try:
        ws = wb["Comprobante"]
        filas_depurado = list(ws.iter_rows(values_only=True))[1:]
    finally:
        wb.close()

    assert len(filas_depurado) == 2
    for fila in filas_depurado:
        assert any(v is not None and v != "" for v in fila)
