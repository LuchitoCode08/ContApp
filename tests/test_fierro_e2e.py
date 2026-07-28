"""Test end-to-end del proceso Fierro.

Genera un Excel sintetico con la hoja 'Diario 2026', ejecuta
ProcesoFierro y verifica las 2 hojas agregadas.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

RAIZ = Path(__file__).resolve().parent.parent
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from procesos.fierro import ProcesoFierro


# Columnas esperadas por el proceso (orden = orden en el Excel original).
COLS_FIERRO = [
    "Tipo", "Comprobante", "Número", "NIT", "Descripción", "Valor",
    "Fecha", "Fondo", "Centro de costos", "Cuenta", "Programa", "D/C",
    "Base Retención", "Tip. cruce", "Com. cruce", "Nro. cruce",
    "Nombre de la cuenta", "Nombre de la entidad", "Desc.Asiento",
]


def _crear_excel_sintetico(ruta: Path, filas: list[list] | None = None) -> None:
    """Crea un Excel con la hoja 'Diario 2026' y filas basicas.

    Por defecto 2 filas neutras (Tipo NM, cuenta 111111) que NO caen
    en ninguno de los filtros (C&C, Ventas FC, Consignaciones BD,
    Compras, Diferencia CC, OP/NC, SE/SG/ST, TC) para que pasen
    tal cual al comprobante.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Diario 2026"
    ws.append(COLS_FIERRO)
    if filas is None:
        filas = [
            ["NM", "001", "1", "123456", "Pago X", "1000.00",
             "01/01/2026", "FOPNAL", "CC01", "111111", "01", "D",
             "0", "", "", "", "Cuenta X", "Entidad X", "Asiento X"],
            ["NM", "001", "2", "123456", "Pago Y", "-1000.00",
             "02/01/2026", "FOPNAL", "CC01", "111111", "01", "C",
             "0", "", "", "", "Cuenta Y", "Entidad Y", "Asiento Y"],
        ]
    for fila in filas:
        ws.append(fila)
    wb.save(ruta)


# --------------------------------------------------------------------
# Validacion
# --------------------------------------------------------------------

def test_validar_archivos_excel_valido(tmp_path: Path) -> None:
    """Excel valido no debe producir error."""
    excel = tmp_path / "fierro.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoFierro()
    assert proceso.validar_archivos([excel]) is None


def test_validar_archivos_sin_archivos() -> None:
    """Lista vacia -> error."""
    proceso = ProcesoFierro()
    error = proceso.validar_archivos([])
    assert error is not None
    assert "al menos un archivo" in error.lower()


def test_validar_archivos_demasiados(tmp_path: Path) -> None:
    """Mas de 1 archivo -> error."""
    a1 = tmp_path / "uno.xlsx"
    a2 = tmp_path / "dos.xlsx"
    _crear_excel_sintetico(a1)
    _crear_excel_sintetico(a2)
    proceso = ProcesoFierro()
    error = proceso.validar_archivos([a1, a2])
    assert error is not None
    assert "un solo" in error.lower()


def test_validar_archivos_extension_invalida(tmp_path: Path) -> None:
    """Archivo .csv (no Excel) -> error."""
    csv = tmp_path / "falso.csv"
    csv.write_text("datos")
    proceso = ProcesoFierro()
    error = proceso.validar_archivos([csv])
    assert error is not None
    assert "no es un excel" in error.lower()


# --------------------------------------------------------------------
# Ejecucion en modo prueba (no toca el original)
# --------------------------------------------------------------------

def test_ejecutar_modo_prueba_genera_archivo_en_carpeta_prueba(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """En modo_prueba, el Excel procesado debe quedar en
    resultados/<proceso>/_prueba_YYYY-MM/ con sufijo _prueba."""
    import procesos.fierro as modulo

    # ProcesoFierro.__init__ lee RAIZ para resolver la ruta de los
    # JSONs (jsons/fierro/). Copiamos los JSONs reales a tmp_path
    # para que el proceso pueda inicializarse sin tocar el repo.
    jsons_fierro_destino = tmp_path / "jsons" / "fierro"
    jsons_fierro_destino.mkdir(parents=True)
    for nombre in ("mapeo_auxiliares.json", "mapeo_descripciones.json",
                   "mapeo_tarjetas.json"):
        shutil.copy2(RAIZ / "jsons" / "fierro" / nombre,
                     jsons_fierro_destino / nombre)

    monkeypatch.setattr(modulo, "RAIZ", tmp_path)
    monkeypatch.setattr(modulo, "RESULTADOS_DIR", tmp_path)
    (tmp_path / "resultados").mkdir()

    excel = tmp_path / "Interfaz.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoFierro()
    resultado = proceso.ejecutar([excel], modo_prueba=True)

    assert resultado.exito, resultado.mensaje
    assert len(resultado.archivos_salida) == 1
    salida = resultado.archivos_salida[0]
    assert salida.exists()
    # En Fierro, en modo_prueba NO se modifica in-place: el archivo
    # se copia a una carpeta "_prueba_YYYY-MM/" bajo resultados/.
    assert any(part.startswith("_prueba_") for part in salida.parts)


# --------------------------------------------------------------------
# Ejecucion en modo produccion (modifica el original)
# --------------------------------------------------------------------

def test_ejecutar_modo_produccion_modifica_original(tmp_path: Path) -> None:
    """En modo produccion, el Excel original debe modificarse in-place."""
    excel = tmp_path / "Interfaz.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoFierro()
    resultado = proceso.ejecutar([excel], modo_prueba=False)

    assert resultado.exito, resultado.mensaje
    assert resultado.archivos_salida[0] == excel
    assert resultado.archivos_salida_originales == [excel]
    # El archivo debe seguir existiendo.
    assert excel.exists()


# --------------------------------------------------------------------
# Estructura del Excel resultante
# --------------------------------------------------------------------

def test_excel_resultado_tiene_2_hojas_esperadas(tmp_path: Path) -> None:
    """El Excel final debe contener la hoja original 'Diario 2026'
    (preservada) mas las 2 hojas nuevas 'Diario 2026 - Copia' y
    'Comprobante'."""
    from openpyxl import load_workbook

    excel = tmp_path / "Interfaz.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoFierro()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        nombres = set(wb.sheetnames)
    finally:
        wb.close()
    assert "Diario 2026" in nombres
    assert "Diario 2026 - Copia" in nombres
    assert "Comprobante" in nombres


def test_hoja_comprobante_tiene_mismas_filas_que_copia(tmp_path: Path) -> None:
    """Para filas neutras (Tipo NM, cuenta 111111), el comprobante debe
    tener las mismas filas que la copia."""
    from openpyxl import load_workbook

    excel = tmp_path / "Interfaz.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoFierro()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        ws_copia = wb["Diario 2026 - Copia"]
        n_copia = sum(1 for _ in ws_copia.iter_rows()) - 1  # restar header
        ws_comp = wb["Comprobante"]
        n_comp = sum(1 for _ in ws_comp.iter_rows()) - 1
    finally:
        wb.close()
    assert n_copia == 2
    assert n_comp == 2


# --------------------------------------------------------------------
# Detalles del ResultadoProceso
# --------------------------------------------------------------------

def test_resultado_incluye_detalles_de_filas(tmp_path: Path) -> None:
    """El resultado debe incluir detalles con conteo de filas."""
    excel = tmp_path / "Interfaz.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoFierro()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito
    assert resultado.detalles is not None
    assert resultado.detalles["filas_originales"] == 2
    assert resultado.detalles["filas_comprobante"] == 2


def test_log_prefix_es_el_de_fierro() -> None:
    """LOG_PREFIX debe identificar a Fierro."""
    assert ProcesoFierro().LOG_PREFIX == "[Fierro]"


# --------------------------------------------------------------------
# Manejo de errores
# --------------------------------------------------------------------

def test_ejecutar_hoja_inexistente_retorna_error(tmp_path: Path) -> None:
    """Si el Excel no tiene la hoja 'Diario 2026', debe fallar elegante."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OtraHoja"
    ws.append(["a", "b"])
    wb.save(tmp_path / "malo.xlsx")

    proceso = ProcesoFierro()
    resultado = proceso.ejecutar([tmp_path / "malo.xlsx"], modo_prueba=False)
    assert not resultado.exito
    assert "diario 2026" in resultado.mensaje.lower()