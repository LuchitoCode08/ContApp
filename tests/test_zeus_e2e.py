"""Test end-to-end del proceso Zeus.

Genera un Excel sintetico con la hoja 'Exportar' (que tenga la
columna Cuenta1), ejecuta ProcesoZeus y verifica las 2 hojas nuevas.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

RAIZ = Path(__file__).resolve().parent.parent
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from procesos.zeus import ProcesoZeus


# Columnas minimas que el proceso necesita en la hoja objetivo.
COLS_ZEUS = [
    "Nit", "Cuenta1", "Tipo_Movto", "valor", "Base", "Fecha", "Concepto",
]


def _crear_excel_sintetico(
    ruta: Path,
    filas: list[list] | None = None,
    nombre_hoja: str = "Exportar",
) -> None:
    """Crea un Excel con una hoja que tenga la columna 'Cuenta1'."""
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    ws.append(COLS_ZEUS)
    if filas is None:
        filas = [
            ["890101681", "11902101", "D", "1500.00", "1500.00",
             "2026-01-15", "Concepto A"],
            ["890101681", "710101", "D", "300.00", "300.00",
             "2026-01-15", "Concepto B"],
        ]
    for fila in filas:
        ws.append(fila)
    wb.save(ruta)


# --------------------------------------------------------------------
# Validacion
# --------------------------------------------------------------------

def test_validar_archivos_excel_valido(tmp_path: Path) -> None:
    excel = tmp_path / "zeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    # Mientras Zeus este en desarrollo, validar_archivos SIEMPRE bloquea.
    # Verificamos que efectivamente devuelve el mensaje de bloqueo.
    if proceso.EN_DESARROLLO:
        error = proceso.validar_archivos([excel])
        assert error is not None
        assert "desarrollo" in error.lower()
    else:
        assert proceso.validar_archivos([excel]) is None


def test_validar_archivos_bloqueado_en_desarrollo(tmp_path: Path) -> None:
    """Si EN_DESARROLLO=True, validar bloquea incluso con un archivo valido."""
    excel = tmp_path / "zeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([excel])
    if proceso.EN_DESARROLLO:
        assert error is not None
        assert proceso.MENSAJE_EN_DESARROLLO in error
    else:
        assert error is None


def test_validar_archivos_sin_archivos() -> None:
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([])
    assert error is not None
    if proceso.EN_DESARROLLO:
        # Mientras este en desarrollo, el mensaje de bloqueo pisa a los demas.
        assert proceso.MENSAJE_EN_DESARROLLO in error
    else:
        assert "al menos un archivo" in error.lower()


def test_validar_archivos_demasiados(tmp_path: Path) -> None:
    a1 = tmp_path / "uno.xlsx"
    a2 = tmp_path / "dos.xlsx"
    _crear_excel_sintetico(a1)
    _crear_excel_sintetico(a2)
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([a1, a2])
    assert error is not None
    if proceso.EN_DESARROLLO:
        assert proceso.MENSAJE_EN_DESARROLLO in error
    else:
        assert "un solo" in error.lower()


def test_validar_archivos_extension_invalida(tmp_path: Path) -> None:
    txt = tmp_path / "falso.txt"
    txt.write_text("datos")
    proceso = ProcesoZeus()
    error = proceso.validar_archivos([txt])
    assert error is not None
    if proceso.EN_DESARROLLO:
        assert proceso.MENSAJE_EN_DESARROLLO in error
    else:
        assert "no es un excel" in error.lower()


# --------------------------------------------------------------------
# Ejecucion en modo prueba
# --------------------------------------------------------------------

def test_ejecutar_modo_prueba_genera_archivo_en_carpeta_prueba(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """En modo_prueba, el Excel se copia a resultados/<proceso>/_prueba_YYYY-MM/
    y el archivo original NO se modifica."""
    import procesos.zeus as modulo
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus esta en desarrollo; ejecutar() bloquea.")

    # ProcesoZeus.__init__ lee RAIZ para resolver el JSON de auxiliares.
    jsons_zeus_destino = tmp_path / "jsons" / "zeus"
    jsons_zeus_destino.mkdir(parents=True)
    shutil.copy2(RAIZ / "jsons" / "zeus" / "auxiliares_zeus.json",
                 jsons_zeus_destino / "auxiliares_zeus.json")

    monkeypatch.setattr(modulo, "RAIZ", tmp_path)
    monkeypatch.setattr(modulo, "RESULTADOS_DIR", tmp_path)
    (tmp_path / "resultados").mkdir()

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=True)

    assert resultado.exito, resultado.mensaje
    assert len(resultado.archivos_salida) == 1
    salida = resultado.archivos_salida[0]
    assert salida.exists()
    assert "zeus" in salida.parts
    assert any(part.startswith("_prueba_") for part in salida.parts)


# --------------------------------------------------------------------
# Ejecucion en modo produccion
# --------------------------------------------------------------------

def test_ejecutar_modo_produccion_modifica_original(tmp_path: Path) -> None:
    """En modo produccion, el Excel original se modifica in-place."""
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus esta en desarrollo; ejecutar() bloquea.")
    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito
    assert resultado.archivos_salida[0] == excel
    assert resultado.archivos_salida_originales == [excel]


# --------------------------------------------------------------------
# Estructura del Excel resultante
# --------------------------------------------------------------------

def test_excel_resultado_tiene_hojas_preservadas_y_nuevas(
    tmp_path: Path,
) -> None:
    """El Excel final debe tener 'Exportar' (preservada) + 2 hojas nuevas."""
    from openpyxl import load_workbook
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus esta en desarrollo; ejecutar() bloquea.")

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        nombres = set(wb.sheetnames)
    finally:
        wb.close()
    assert "Exportar" in nombres
    assert "Exportar - Copia" in nombres
    assert "Comprobante" in nombres


def test_depurado_aplica_auxiliares_8_a_6_digitos(tmp_path: Path) -> None:
    """Cuenta1 '11902101' debe quedar como '119021' (auxiliar de Zeus)."""
    from openpyxl import load_workbook
    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus esta en desarrollo; ejecutar() bloquea.")

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        ws = wb["Exportar - Copia"]
        filas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header = filas[0]
    idx_cuenta1 = header.index("Cuenta1")
    cuentas = [fila[idx_cuenta1] for fila in filas[1:]]
    # La primera fila era '11902101' y debe quedar como '119021'.
    assert "119021" in cuentas


def test_comprobante_agrega_columnas_nuevas(tmp_path: Path) -> None:
    """La hoja 'Comprobante' debe tener Valor2, BaseAbs, Tarifa."""
    from openpyxl import load_workbook

    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito

    wb = load_workbook(excel, read_only=True)
    try:
        ws = wb["Comprobante"]
        header = next(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    assert "Valor2" in header
    assert "BaseAbs" in header
    assert "Tarifa" in header


# --------------------------------------------------------------------
# Detalles del ResultadoProceso
# --------------------------------------------------------------------

def test_resultado_incluye_detalles_de_filas(tmp_path: Path) -> None:
    excel = tmp_path / "InterfazZeus.xlsx"
    _crear_excel_sintetico(excel)
    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito
    assert resultado.detalles is not None
    assert resultado.detalles["filas_originales"] == 2
    assert resultado.detalles["filas_comprobante"] == 2


def test_log_prefix_es_el_de_zeus() -> None:
    assert ProcesoZeus().LOG_PREFIX == "[Zeus]"


# --------------------------------------------------------------------
# Errores
# --------------------------------------------------------------------

def test_ejecutar_excel_sin_hoja_exportar_retorna_error(tmp_path: Path) -> None:
    """Si el Excel no tiene la hoja 'Exportar', debe fallar con mensaje claro."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SinCuenta1"
    ws.append(["a", "b"])
    wb.save(tmp_path / "malo.xlsx")

    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([tmp_path / "malo.xlsx"], modo_prueba=False)
    assert not resultado.exito
    assert "exportar" in resultado.mensaje.lower()

# --------------------------------------------------------------------
# Agrupacion sin filas vacias
# --------------------------------------------------------------------

def test_agrupacion_no_genera_filas_vacias(tmp_path: Path) -> None:
    """Al agrupar por Nit/Cuenta1/Fecha no deben quedar filas NaN."""
    from openpyxl import load_workbook

    if ProcesoZeus.EN_DESARROLLO:
        pytest.skip("Zeus esta en desarrollo; ejecutar() bloquea.")

    filas = [
        # Dos filas agrupables: mismo Nit, Cuenta1 (de 4 digitos) y Fecha.
        ["890101681", "7101", "D", "100.00", "100.00",
         "2026-01-15", "Concepto A"],
        ["890101681", "7101", "D", "200.00", "200.00",
         "2026-01-15", "Concepto B"],
        # Una fila no agrupable para conservar mezcla (auxiliar a 6 digitos).
        ["890101681", "11902101", "D", "1500.00", "1500.00",
         "2026-01-15", "Concepto C"],
    ]
    excel = tmp_path / "ZeusAgrupado.xlsx"
    _crear_excel_sintetico(excel, filas=filas)

    proceso = ProcesoZeus()
    resultado = proceso.ejecutar([excel], modo_prueba=False)
    assert resultado.exito, resultado.mensaje

    # La hoja Comprobante debe tener exactamente 2 filas (grupo colapsado + no agrupable).
    wb = load_workbook(excel, read_only=True)
    try:
        ws = wb["Comprobante"]
        filas_depurado = list(ws.iter_rows(values_only=True))[1:]
    finally:
        wb.close()

    assert len(filas_depurado) == 2
    # Ninguna fila debe estar completamente vacia.
    for fila in filas_depurado:
        assert any(v is not None and v != "" for v in fila)
