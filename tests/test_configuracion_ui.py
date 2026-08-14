"""Tests pytest-qt de ``PantallaConfiguracion``.

Cubre:
    - Estado inicial: widgets correctos, 6 columnas en la tabla, combo boxes
      con items esperados, total de registros == 0 con log vacio.
    - Carga de registros: ``refrescar()`` lee el log y llena la tabla.
    - Filtros: cada uno de los 5 filtros (fecha, proceso, nivel, modo, texto)
      reduce correctamente la cantidad de filas visibles.
    - Buscador con debounce de 300 ms: cambiar el texto dispara el filtro
      despues del delay (NO en cada tecla).
    - Refresco silencioso periodico cada 5 s: detecta cuando hay registros
      nuevos en disco.
    - Refresco NO altera el total cuando no hay cambios.
    - Exportar a CSV genera un archivo con header + filas correctas.
    - Exportar a Excel genera un .xlsx valido con 6 columnas.
    - Limpieza de registros antiguos filtra por fecha.

Estrategia:
    - offscreen Qt.
    - Logs sinteticos en ``tmp_path``.
    - Monkey-patching de ``BITACORA_LOG`` y ``leer_registros`` (segun
      el caso: o el archivo o el retorno directo).
    - Monkey-patching de ``QMessageBox`` / ``QFileDialog`` para no abrir
      dialogs reales.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

# Forzar offscreen ANTES de cualquier import Qt.
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import pytest  # noqa: E402
from PySide6.QtCore import QDate, Qt  # noqa: E402
from PySide6.QtWidgets import QApplication, QFileDialog, QInputDialog, QMessageBox  # noqa: E402

RAIZ_PROY = Path(__file__).resolve().parent.parent
if str(RAIZ_PROY) not in sys.path:
    sys.path.insert(0, str(RAIZ_PROY))


@pytest.fixture(scope="module")
def qt_app() -> QApplication:
    """Una sola QApplication por modulo (Qt no permite mas de una)."""
    app = QApplication.instance() or QApplication(sys.argv)
    yield app


# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------

def _escribir_log(path: Path, lineas: list[str]) -> None:
    """Escribe lineas de log en formato estandar (cada una con \\n)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lineas) + "\n", encoding="utf-8")


def _linea_log(
    fecha: str, nivel: str, mensaje: str, modulo: str = "contapp",
) -> str:
    """Genera una linea de log sintetica con formato compatible con el parser."""
    return f"{fecha} [{nivel}] {modulo}: {mensaje}"


def _patch_log(monkeypatch, log_path: Path) -> None:
    """Redirige tanto ``cfg_mod.BITACORA_LOG`` como ``app.config.BITACORA_LOG``.

    El modulo importa la constante al cargar (``from app.config import
    BITACORA_LOG``), asi que parchear solo el modulo no es suficiente
    si ``leer_registros`` resuelve la ruta desde ``app.config``.
    """
    import app.config as app_cfg
    import ui.ventanas.configuracion as cfg_mod
    monkeypatch.setattr(app_cfg, "BITACORA_LOG", log_path)
    monkeypatch.setattr(cfg_mod, "BITACORA_LOG", log_path)


def _make_pantalla(qt_app, tmp_path, monkeypatch) -> "PantallaConfiguracion":
    """Helper que crea una ``PantallaConfiguracion`` apuntando a tmp_path."""
    log_path = tmp_path / "bitacora.log"
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    # Detenemos el timer periodico para que no interfiera en tests.
    p._timer.stop()
    return p


# ============================================================
# Estado inicial
# ============================================================

def test_pantalla_inicial_tiene_6_columnas(qt_app, tmp_path, monkeypatch) -> None:
    """La tabla debe tener 6 columnas: Fecha, Proceso, Nivel, Modo, Modulo, Mensaje."""
    p = _make_pantalla(qt_app, tmp_path, monkeypatch)
    try:
        assert p._tabla.columnCount() == 6
        headers = [
            p._tabla.horizontalHeaderItem(i).text()
            for i in range(p._tabla.columnCount())
        ]
        assert headers == ["Fecha", "Proceso", "Nivel", "Modo", "Modulo", "Mensaje"]
    finally:
        p.deleteLater()


def test_pantalla_inicial_combos_tienen_todos_predeterminado(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Los 3 filtros (Proceso, Nivel, Modo) arrancan en 'Todos'."""
    p = _make_pantalla(qt_app, tmp_path, monkeypatch)
    try:
        assert p._filtro_proceso.currentText() == "Todos"
        assert p._filtro_nivel.currentText() == "Todos"
        assert p._filtro_modo.currentText() == "Todos"
        # Y los items son los esperados.
        assert p._filtro_proceso.count() == 4   # Todos + 3 procesos
        assert p._filtro_nivel.count() == 5     # Todos + 4 niveles
        assert p._filtro_modo.count() == 3      # Todos + Produccion + Prueba
    finally:
        p.deleteLater()


def test_pantalla_con_log_vacio_muestra_0_registros(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Si el log no existe o esta vacio, la tabla tiene 0 filas."""
    p = _make_pantalla(qt_app, tmp_path, monkeypatch)
    try:
        assert p._tabla.rowCount() == 0
        assert "0" in p._lbl_total.text()
    finally:
        p.deleteLater()


def test_pantalla_arranca_con_timer_de_refresco_detenido(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """El timer periodico de 5s arranca detenido (lo paramos en el fixture)."""
    p = _make_pantalla(qt_app, tmp_path, monkeypatch)
    try:
        assert not p._timer.isActive()
    finally:
        p.deleteLater()


# ============================================================
# Carga de registros
# ============================================================

def test_refrescar_carga_registros_del_log(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """``refrescar()`` lee el log del disco y llena la tabla."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] Generado: foo.xlsx"),
        _linea_log("2026-07-30 10:01:00", "WARNING", "[Fierro] Error leve"),
        _linea_log("2026-07-30 10:02:00", "ERROR", "[Zeus] Fallo critico"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        assert p._tabla.rowCount() == 3
        assert "3" in p._lbl_total.text()
    finally:
        p.deleteLater()


# ============================================================
# Filtros: Proceso
# ============================================================

def test_filtro_proceso_reduce_filas(qt_app, tmp_path, monkeypatch) -> None:
    """Seleccionar un proceso especifico deja solo sus registros."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
        _linea_log("2026-07-30 10:01:00", "INFO", "[Fierro] b"),
        _linea_log("2026-07-30 10:02:00", "INFO", "[Comprobante] c"),
        _linea_log("2026-07-30 10:03:00", "INFO", "[Zeus] d"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        # Por defecto: 4 filas.
        assert p._tabla.rowCount() == 4

        # Filtrar por Comprobante -> 2 filas.
        p._filtro_proceso.setCurrentText("Comprobante")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 2

        # Filtrar por Zeus -> 1 fila.
        p._filtro_proceso.setCurrentText("Zeus")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 1
    finally:
        p.deleteLater()


# ============================================================
# Filtros: Nivel
# ============================================================

def test_filtro_nivel_reduce_filas(qt_app, tmp_path, monkeypatch) -> None:
    """Filtrar por nivel deja solo los registros de ese nivel."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
        _linea_log("2026-07-30 10:01:00", "WARNING", "[Fierro] b"),
        _linea_log("2026-07-30 10:02:00", "ERROR", "[Comprobante] c"),
        _linea_log("2026-07-30 10:03:00", "INFO", "[Fierro] d"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        assert p._tabla.rowCount() == 4

        p._filtro_nivel.setCurrentText("INFO")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 2

        p._filtro_nivel.setCurrentText("ERROR")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 1
    finally:
        p.deleteLater()


# ============================================================
# Filtros: Modo (PRUEBA / Produccion)
# ============================================================

def test_filtro_modo_prueba(qt_app, tmp_path, monkeypatch) -> None:
    """Filtrar por 'Prueba' deja solo registros con marca [PRUEBA]."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] prod [PRUEBA]"),
        _linea_log("2026-07-30 10:01:00", "INFO", "[Fierro] sin marca"),
        _linea_log("2026-07-30 10:02:00", "INFO", "[Comprobante] c [PRUEBA]"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        assert p._tabla.rowCount() == 3

        p._filtro_modo.setCurrentText("Prueba")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 2
    finally:
        p.deleteLater()


def test_filtro_modo_produccion(qt_app, tmp_path, monkeypatch) -> None:
    """Filtrar por 'Produccion' deja solo registros SIN marca [PRUEBA]."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] prod [PRUEBA]"),
        _linea_log("2026-07-30 10:01:00", "INFO", "[Fierro] sin marca"),
        _linea_log("2026-07-30 10:02:00", "INFO", "[Comprobante] c [PRUEBA]"),
        _linea_log("2026-07-30 10:03:00", "INFO", "[Zeus] d"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        p._filtro_modo.setCurrentText("Produccion")
        p._aplicar_filtros()
        QApplication.processEvents()
        # Solo los 2 sin marca [PRUEBA].
        assert p._tabla.rowCount() == 2
    finally:
        p.deleteLater()


# ============================================================
# Filtros: Buscador con debounce
# ============================================================

def test_buscador_filtra_por_texto(qt_app, tmp_path, monkeypatch) -> None:
    """El buscador reduce filas segun el texto contenido en el mensaje."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] Excel procesado: foo.xlsx"),
        _linea_log("2026-07-30 10:01:00", "INFO", "[Fierro] Planilla generada"),
        _linea_log("2026-07-30 10:02:00", "INFO", "[Comprobante] Excel procesado: bar.xlsx"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        assert p._tabla.rowCount() == 3

        # Tipear "excel" -> quedan 2.
        p._buscador.setText("excel")
        # Llamamos _aplicar_filtros directo para evitar esperar el debounce.
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 2

        # Busqueda case-insensitive.
        p._buscador.setText("PLANILLA")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 1
    finally:
        p.deleteLater()


def test_buscador_tiene_debounce_de_300ms(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """El timer del buscador es single-shot y tiene interval de 300 ms."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        assert p._buscador_timer.isSingleShot() is True
        assert p._buscador_timer.interval() == 300
    finally:
        p.deleteLater()


# ============================================================
# Filtros: Fecha
# ============================================================

def test_filtro_fecha_reduce_filas(qt_app, tmp_path, monkeypatch) -> None:
    """Filtrar por un rango de fechas excluye registros fuera del rango."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-25 10:00:00", "INFO", "[Comprobante] viejo 1"),
        _linea_log("2026-07-28 10:00:00", "INFO", "[Comprobante] reciente 1"),
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] reciente 2"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        # Por defecto: 3 filas (rango = hoy - 1 mes ... hoy).
        assert p._tabla.rowCount() == 3

        # Restringir a 2026-07-28 ... 2026-07-30 -> 2 filas.
        p._fecha_desde.setDate(QDate(2026, 7, 28))
        p._fecha_hasta.setDate(QDate(2026, 7, 30))
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 2
    finally:
        p.deleteLater()


# ============================================================
# Refresco silencioso
# ============================================================

def test_refresco_silencioso_detecta_nuevos_registros(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Si hay registros nuevos en disco, el refresco los detecta."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        assert p._tabla.rowCount() == 1

        # Simulamos que alguien escribio una linea mas en el log.
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(_linea_log("2026-07-30 10:05:00", "INFO", "[Fierro] b") + "\n")

        # Disparamos el refresco manualmente.
        p._refresco_silencioso()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 2
    finally:
        p.deleteLater()


def test_refresco_silencioso_no_altera_sin_cambios(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Si el log no cambio, ``_refresco_silencioso`` no altera el total."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        total_antes = p._tabla.rowCount()
        assert total_antes == 1

        # Sin cambios -> el total sigue igual.
        p._refresco_silencioso()
        QApplication.processEvents()
        assert p._tabla.rowCount() == total_antes
    finally:
        p.deleteLater()


# ============================================================
# Boton Aplicar filtros
# ============================================================

def test_btn_aplicar_dispara_filtros(qt_app, tmp_path, monkeypatch) -> None:
    """Click en 'Aplicar filtros' invoca ``_aplicar_filtros``."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
        _linea_log("2026-07-30 10:01:00", "INFO", "[Fierro] b"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        # Spy sobre _aplicar_filtros.
        llamado = {"count": 0}
        original = p._aplicar_filtros

        def spy():
            llamado["count"] += 1
            return original()

        p._aplicar_filtros = spy
        p.btn_aplicar.click()
        QApplication.processEvents()
        assert llamado["count"] == 1
    finally:
        p.deleteLater()


# ============================================================
# Exportar CSV
# ============================================================

def test_exportar_csv_genera_archivo_con_header_y_filas(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Exportar a CSV escribe un archivo con header + las filas filtradas."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
        _linea_log("2026-07-30 10:01:00", "INFO", "[Fierro] b"),
    ])
    _patch_log(monkeypatch, log_path)
    # Silenciar QMessageBox + patchear QFileDialog para devolver tmp_path.
    monkeypatch.setattr(QMessageBox, "information", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Ok,
    ))
    monkeypatch.setattr(QMessageBox, "critical", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Ok,
    ))
    csv_path = tmp_path / "export.csv"
    monkeypatch.setattr(QFileDialog, "getSaveFileName", staticmethod(
        lambda *a, **k: (str(csv_path), "CSV (*.csv)"),
    ))

    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        p._exportar_csv()
        QApplication.processEvents()
        # El archivo existe y tiene al menos 1 linea (header).
        assert csv_path.exists()
        contenido = csv_path.read_text(encoding="utf-8")
        lineas = contenido.strip().split("\n")
        # header + 2 filas = 3 lineas.
        assert len(lineas) == 3
        # El header tiene las 6 columnas.
        header = lineas[0]
        for col in ("Fecha", "Proceso", "Nivel", "Modo", "Modulo", "Mensaje"):
            assert col in header
    finally:
        p.deleteLater()


# ============================================================
# Exportar Excel
# ============================================================

def test_exportar_excel_genera_archivo_valido(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Exportar a Excel genera un .xlsx valido con 6 columnas."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] a"),
        _linea_log("2026-07-30 10:01:00", "WARNING", "[Fierro] b"),
    ])
    _patch_log(monkeypatch, log_path)
    monkeypatch.setattr(QMessageBox, "information", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Ok,
    ))
    monkeypatch.setattr(QMessageBox, "critical", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Ok,
    ))
    xlsx_path = tmp_path / "export.xlsx"
    monkeypatch.setattr(QFileDialog, "getSaveFileName", staticmethod(
        lambda *a, **k: (str(xlsx_path), "Excel (*.xlsx)"),
    ))

    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        p._exportar_excel()
        QApplication.processEvents()
        assert xlsx_path.exists()
        # Validar estructura del .xlsx.
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # Header + 2 filas = 3 filas.
        assert ws.max_row == 3
        assert ws.max_column == 6
        # Header de la primera fila.
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["Fecha", "Proceso", "Nivel", "Modo", "Modulo", "Mensaje"]
    finally:
        p.deleteLater()


# ============================================================
# Limpieza de registros antiguos
# ============================================================

def test_limpiar_antiguos_filtra_por_dias(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """``_limpiar_antiguos`` elimina lineas mas viejas que N dias."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-06-01 10:00:00", "INFO", "[Comprobante] muy viejo"),
        _linea_log("2026-07-28 10:00:00", "INFO", "[Fierro] reciente"),
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] reciente 2"),
    ])
    _patch_log(monkeypatch, log_path)
    monkeypatch.setattr(QMessageBox, "warning", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Yes,
    ))
    monkeypatch.setattr(QMessageBox, "information", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Ok,
    ))
    monkeypatch.setattr(QMessageBox, "critical", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Ok,
    ))
    # Pedimos 14 dias: se elimina solo el de junio.
    monkeypatch.setattr(QInputDialog, "getInt", staticmethod(
        lambda *a, **k: (14, True),
    ))

    from datetime import datetime as dt_real
    import ui.ventanas.configuracion as cfg_mod
    # Congelamos la fecha actual para que el test no dependa del dia de ejecucion.
    class _FechaFija:
        @staticmethod
        def now() -> dt_real:
            return dt_real(2026, 8, 1, 12, 0, 0)
        @staticmethod
        def strptime(*args, **kwargs):
            return dt_real.strptime(*args, **kwargs)
        @staticmethod
        def combine(*args, **kwargs):
            return dt_real.combine(*args, **kwargs)
    monkeypatch.setattr(cfg_mod, "datetime", _FechaFija())

    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        # Llamamos directo al handler del boton Limpiar.
        p._limpiar_antiguos()
        QApplication.processEvents()
        # El log ahora NO debe tener la linea del 2026-06-01.
        contenido = log_path.read_text(encoding="utf-8")
        assert "muy viejo" not in contenido
        assert "reciente" in contenido
    finally:
        p.deleteLater()


def test_limpiar_antiguos_cancela_si_usuario_dice_no(
    qt_app, tmp_path, monkeypatch,
) -> None:
    """Si el usuario cancela el QInputDialog, no se elimina nada."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    lineas_originales = [
        _linea_log("2026-06-01 10:00:00", "INFO", "[Comprobante] muy viejo"),
        _linea_log("2026-07-30 10:00:00", "INFO", "[Fierro] reciente"),
    ]
    _escribir_log(log_path, lineas_originales)
    _patch_log(monkeypatch, log_path)
    monkeypatch.setattr(QMessageBox, "warning", staticmethod(
        lambda *a, **k: QMessageBox.StandardButton.Yes,
    ))
    # getInt retorna (value, ok=False) -> el usuario cancelo.
    monkeypatch.setattr(QInputDialog, "getInt", staticmethod(
        lambda *a, **k: (14, False),
    ))

    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        p._limpiar_antiguos()
        QApplication.processEvents()
        # El archivo NO se toco.
        contenido = log_path.read_text(encoding="utf-8")
        assert "muy viejo" in contenido
        assert "reciente" in contenido
    finally:
        p.deleteLater()


# ============================================================
# Coherencia: filtros combinables
# ============================================================

def test_filtros_combinados(qt_app, tmp_path, monkeypatch) -> None:
    """Proceso + Nivel + Buscador funcionan en conjunto (AND logico)."""
    import ui.ventanas.configuracion as cfg_mod
    log_path = tmp_path / "bitacora.log"
    _escribir_log(log_path, [
        _linea_log("2026-07-30 10:00:00", "INFO", "[Comprobante] Excel procesado: a.xlsx"),
        _linea_log("2026-07-30 10:01:00", "WARNING", "[Comprobante] Fallo en Excel b.xlsx"),
        _linea_log("2026-07-30 10:02:00", "INFO", "[Fierro] Excel procesado: c.xlsx"),
        _linea_log("2026-07-30 10:03:00", "INFO", "[Comprobante] Planilla d.xlsx"),
    ])
    _patch_log(monkeypatch, log_path)
    from ui.ventanas.configuracion import PantallaConfiguracion
    p = PantallaConfiguracion()
    p._timer.stop()
    try:
        # Comprobante + INFO + "excel" -> 1 fila ("Excel procesado: a.xlsx").
        p._filtro_proceso.setCurrentText("Comprobante")
        p._filtro_nivel.setCurrentText("INFO")
        p._buscador.setText("excel")
        p._aplicar_filtros()
        QApplication.processEvents()
        assert p._tabla.rowCount() == 1
        # Verificar que es el registro correcto.
        mensaje = p._tabla.item(0, 5).text()
        assert "a.xlsx" in mensaje
    finally:
        p.deleteLater()
