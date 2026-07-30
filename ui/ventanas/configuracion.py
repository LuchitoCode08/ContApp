"""Pantalla Configuracion: visor de la bitacora con filtros y exportacion.

Permite ver, filtrar, exportar y limpiar los registros generados por la app.
"""
from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PySide6.QtCore import QDate, Qt, QTimer
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.config import BITACORA_LOG, get_config
from utils.bitacora import (
    _extraer_proceso,
    es_modo_prueba,
    leer_registros,
    log,
    quitar_marca_prueba,
)


# Mapeo de nivel a color de fondo (para resaltar).
# Los colores "INFO" y "DEBUG" dependen del tema actual y se computan en
# ``_color_nivel()`` para que la tabla se vea bien tanto en claro como en
# oscuro.
COLOR_NIVEL_BASE = {
    "ERROR": "#FDAAB3",    # rojo suave
    "WARNING": "#F3C47E",  # naranja suave
    "INFO": None,          # se calcula en _color_nivel()
    "DEBUG": None,         # se calcula en _color_nivel()
}


def _color_nivel(nivel: str) -> str:
    """Devuelve el color de fondo para una fila segun nivel y tema actual."""
    from ui.recursos.tema import _paleta
    p = _paleta()
    if nivel == "INFO":
        return p.surface
    if nivel == "DEBUG":
        return p.surface_alt
    return COLOR_NIVEL_BASE.get(nivel, p.surface)

PROCESOS = ["Todos", "Comprobante", "Fierro", "Zeus"]
NIVELES = ["Todos", "INFO", "WARNING", "ERROR", "DEBUG"]
MODOS = ["Todos", "Produccion", "Prueba"]


class PantallaConfiguracion(QWidget):
    """Visor de la bitacora con filtros y exportacion."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cfg = get_config()
        self._registros: list[dict] = []
        self._construir_ui()
        self.refrescar()

        # Refresco automatico cada 5 segundos (para ver logs en vivo).
        self._timer = QTimer(self)
        self._timer.setInterval(5000)
        self._timer.timeout.connect(self._refresco_silencioso)
        self._timer.start()

    # -- UI ---------------------------------------------------------

    def _construir_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(14)

        titulo = QLabel("📋  Bitácora de actividad")
        titulo.setStyleSheet(
            "font-size: 22px; font-weight: 700; padding-bottom: 4px;"
        )
        layout.addWidget(titulo)

        # --- Filtros ------------------------------------------------
        self._filtros = QFrame()
        self._filtros.setObjectName("filtros_frame")
        self._filtros.setFrameShape(QFrame.Shape.StyledPanel)
        fl = QHBoxLayout(self._filtros)
        fl.setContentsMargins(8, 8, 8, 8)

        fl.addWidget(QLabel("Desde:"))
        self._fecha_desde = QDateEdit()
        self._fecha_desde.setCalendarPopup(True)
        self._fecha_desde.setDisplayFormat("yyyy-MM-dd")
        self._fecha_desde.setDate(QDate.currentDate().addMonths(-1))
        fl.addWidget(self._fecha_desde)

        fl.addWidget(QLabel("Hasta:"))
        self._fecha_hasta = QDateEdit()
        self._fecha_hasta.setCalendarPopup(True)
        self._fecha_hasta.setDisplayFormat("yyyy-MM-dd")
        self._fecha_hasta.setDate(QDate.currentDate())
        fl.addWidget(self._fecha_hasta)

        fl.addSpacing(12)
        fl.addWidget(QLabel("Proceso:"))
        self._filtro_proceso = QComboBox()
        self._filtro_proceso.addItems(PROCESOS)
        fl.addWidget(self._filtro_proceso)

        fl.addSpacing(12)
        fl.addWidget(QLabel("Nivel:"))
        self._filtro_nivel = QComboBox()
        self._filtro_nivel.addItems(NIVELES)
        fl.addWidget(self._filtro_nivel)

        fl.addSpacing(12)
        fl.addWidget(QLabel("Modo:"))
        self._filtro_modo = QComboBox()
        self._filtro_modo.addItems(MODOS)
        fl.addWidget(self._filtro_modo)

        fl.addSpacing(12)
        fl.addWidget(QLabel("Buscar:"))
        self._buscador = QLineEdit()
        self._buscador.setPlaceholderText("texto en el mensaje...")
        self._buscador.setClearButtonEnabled(True)
        # Debounce: aplicamos el filtro 300 ms despues de dejar de tipear.
        self._buscador_timer = QTimer(self)
        self._buscador_timer.setSingleShot(True)
        self._buscador_timer.setInterval(300)
        self._buscador_timer.timeout.connect(self._aplicar_filtros)
        self._buscador.textChanged.connect(
            lambda _: self._buscador_timer.start()
        )
        fl.addWidget(self._buscador, 1)

        # Botones de filtro.
        self.btn_aplicar = QPushButton("Aplicar filtros")
        self.btn_aplicar.clicked.connect(self._aplicar_filtros)
        fl.addWidget(self.btn_aplicar)

        self.btn_refrescar = QPushButton("Refrescar")
        self.btn_refrescar.clicked.connect(self.refrescar)
        fl.addWidget(self.btn_refrescar)

        layout.addWidget(self._filtros)
        self._aplicar_tema(self._tema_actual())

        # --- Tabla --------------------------------------------------
        self._tabla = QTableWidget()
        self._tabla.setColumnCount(6)
        self._tabla.setHorizontalHeaderLabels(
            ["Fecha", "Proceso", "Nivel", "Modo", "Modulo", "Mensaje"]
        )
        self._tabla.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tabla.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self._tabla.setAlternatingRowColors(True)
        hdr = self._tabla.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self._tabla.verticalHeader().setVisible(False)
        layout.addWidget(self._tabla, 1)

        # --- Footer -------------------------------------------------
        footer = QHBoxLayout()
        self._lbl_total = QLabel("  0 registros")
        footer.addWidget(self._lbl_total)
        footer.addStretch()

        self.btn_exportar_excel = QPushButton("📊  Exportar Excel")
        self.btn_exportar_excel.clicked.connect(self._exportar_excel)
        footer.addWidget(self.btn_exportar_excel)

        self.btn_exportar_csv = QPushButton("📄  Exportar CSV")
        self.btn_exportar_csv.clicked.connect(self._exportar_csv)
        footer.addWidget(self.btn_exportar_csv)

        self.btn_limpiar = QPushButton("🗑  Limpiar antiguos")
        self.btn_limpiar.setObjectName("danger")
        self.btn_limpiar.clicked.connect(self._limpiar_antiguos)
        footer.addWidget(self.btn_limpiar)

        layout.addLayout(footer)

        self._aplicar_tema(self._tema_actual())

    def _tema_actual(self):
        from ui.recursos.tema import _paleta
        return _paleta()

    def _aplicar_tema(self, paleta) -> None:
        """Reaplica estilos al cambiar de tema."""
        if hasattr(self, "_filtros") and self._filtros is not None:
            self._filtros.setStyleSheet(
                f"QFrame {{ background-color: {paleta.surface_alt};"
                f" border: 1px solid {paleta.border};"
                " border-radius: 4px; }"
            )
        if hasattr(self, "_lbl_total") and self._lbl_total is not None:
            self._lbl_total.setStyleSheet(
                f"color: {paleta.fg_muted}; font-size: 12px;"
            )
        if hasattr(self, "btn_limpiar") and self.btn_limpiar is not None:
            # Estilo explicito con fondo transparente para evitar que el
            # background-color del QSS global #danger (rojo intenso) tape
            # el texto. Asi solo se ve el borde + texto rojo en estado
            # normal, y al hover el fondo se vuelve gris claro.
            self.btn_limpiar.setStyleSheet(
                f"QPushButton {{"
                f" color: {paleta.danger};"
                f" background-color: transparent;"
                f" border: 1px solid {paleta.danger};"
                f" border-radius: 4px;"
                f" padding: 6px 12px;"
                f" font-weight: 600;"
                f" }}"
                f"QPushButton:hover {{"
                f" background-color: {paleta.surface_alt};"
                f" }}"
                f"QPushButton:pressed {{"
                f" background-color: {paleta.border};"
                f" }}"
                f"QPushButton:disabled {{"
                f" color: {paleta.fg_disabled};"
                f" border-color: {paleta.border};"
                f" }}"
            )
        # Repintar SOLO los colores de las celdas existentes (no
        # re-aplicar filtros completos). Esto evita recorrer todos los
        # registros de nuevo, que era el cuello de botella al cambiar
        # de tema (la app se colgaba unos segundos).
        if hasattr(self, "_tabla") and self._tabla.rowCount() > 0:
            self._repintar_colores_tabla()

        # Senales automaticas para algunos filtros.
        self._fecha_desde.dateChanged.connect(
            lambda _: self._buscador_timer.start()
        )
        self._fecha_hasta.dateChanged.connect(
            lambda _: self._buscador_timer.start()
        )
        self._filtro_proceso.currentIndexChanged.connect(
            lambda _: self._buscador_timer.start()
        )
        self._filtro_nivel.currentIndexChanged.connect(
            lambda _: self._buscador_timer.start()
        )
        self._filtro_modo.currentIndexChanged.connect(
            lambda _: self._buscador_timer.start()
        )

    # -- Carga y refresco ------------------------------------------

    def refrescar(self) -> None:
        """Lee la bitacora del disco y aplica los filtros actuales."""
        self._registros = leer_registros()
        self._aplicar_filtros()

    def _refresco_silencioso(self) -> None:
        """Refresco automatico que no pisa el scroll del usuario."""
        nuevos = leer_registros()
        # Solo actualizamos si hay cambios reales.
        if len(nuevos) != len(self._registros) or (
            nuevos and self._registros
            and nuevos[-1]["mensaje_crudo"] != self._registros[-1]["mensaje_crudo"]
        ):
            self._registros = nuevos
            self._aplicar_filtros()

    def _aplicar_filtros(self) -> None:
        if not hasattr(self, "_tabla"):
            return
        fecha_desde = self._fecha_desde.date().toString("yyyy-MM-dd")
        fecha_hasta = self._fecha_hasta.date().toString("yyyy-MM-dd")
        proc = self._filtro_proceso.currentText()
        nivel = self._filtro_nivel.currentText()
        modo = self._filtro_modo.currentText()
        texto = self._buscador.text().strip().lower()

        filtrados = []
        for r in self._registros:
            f = r["fecha"][:10]
            if f < fecha_desde or f > fecha_hasta:
                continue
            if nivel != "Todos" and r["nivel"] != nivel:
                continue
            p = _extraer_proceso(r["mensaje"]).capitalize()
            if proc != "Todos" and p != proc:
                # Si el registro no tiene prefijo de proceso, lo dejamos pasar
                # solo si el filtro es "Todos" (ya cubierto arriba).
                continue
            # Filtro por modo (PRUEBA / PROD / vacio).
            if modo != "Todos":
                prueba = es_modo_prueba(r["mensaje"])
                if modo == "Prueba" and prueba is not True:
                    continue
                if modo == "Produccion" and prueba is not False:
                    continue
            if texto and texto not in r["mensaje"].lower():
                continue
            filtrados.append({**r, "_proceso": p})

        self._llenar_tabla(filtrados)
        self._lbl_total.setText(f"{len(filtrados)} registro(s)")

    def _llenar_tabla(self, registros: list[dict]) -> None:
        from ui.recursos.tema import _paleta
        p = _paleta()
        self._tabla.setRowCount(len(registros))
        for i, r in enumerate(registros):
            # Detectamos modo prueba por la marca "[PRUEBA]" en el mensaje.
            prueba = es_modo_prueba(r["mensaje"])
            mensaje_limpio = quitar_marca_prueba(r["mensaje"])
            modo_str = "PRUEBA" if prueba is True else (
                "PROD" if prueba is False else ""
            )
            items = [
                QTableWidgetItem(r["fecha"]),
                QTableWidgetItem(r.get("_proceso", "")),
                QTableWidgetItem(r["nivel"]),
                QTableWidgetItem(modo_str),
                QTableWidgetItem(r["modulo"]),
                QTableWidgetItem(mensaje_limpio.replace("\n", " | ")),
            ]
            for col, item in enumerate(items):
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                # Forzar foreground explicito para que en oscuro no quede blanco.
                item.setForeground(QColor(p.fg))
                self._tabla.setItem(i, col, item)
            # Color de fondo por nivel (dependiente del tema).
            bg = _color_nivel(r["nivel"])
            for item in items:
                item.setBackground(QColor(bg))
            # Highlight de la columna Modo (PRUEBA en naranja suave).
            if prueba is True:
                items[3].setBackground(QColor("#FFE0B2"))
                items[3].setForeground(QColor("#E65100"))
                font = items[3].font()
                font.setBold(True)
                items[3].setFont(font)
            elif prueba is False:
                items[3].setForeground(QColor("#2E7D32"))

    def _repintar_colores_tabla(self) -> None:
        """Repinta SOLO los colores (bg/fg) de las celdas existentes.

        Es la version rapida que se invoca al cambiar de tema. NO
        recrea los ``QTableWidgetItem`` (eso era el cuello de botella
        en `_llenar_tabla` cuando la tabla tenia cientos de filas).

        Asume que las celdas ya estan pobladas por `_llenar_tabla`
        y que la columna 2 (Nivel) tiene el nivel del registro, lo
        cual usamos para calcular el color de fondo.
        """
        from ui.recursos.tema import _paleta
        p = _paleta()
        filas = self._tabla.rowCount()
        for i in range(filas):
            item_nivel = self._tabla.item(i, 2)
            if item_nivel is None:
                continue
            nivel = item_nivel.text()
            bg = _color_nivel(nivel)
            # Aplicar fg/bg a las 6 columnas.
            for col in range(6):
                item = self._tabla.item(i, col)
                if item is None:
                    continue
                item.setForeground(QColor(p.fg))
                item.setBackground(QColor(bg))
            # Highlight de columna Modo (PRUEBA en naranja, PROD en verde).
            item_modo = self._tabla.item(i, 3)
            if item_modo is None:
                continue
            texto_modo = item_modo.text()
            if texto_modo == "PRUEBA":
                item_modo.setBackground(QColor("#FFE0B2"))
                item_modo.setForeground(QColor("#E65100"))
                font = item_modo.font()
                font.setBold(True)
                item_modo.setFont(font)
            elif texto_modo == "PROD":
                item_modo.setForeground(QColor("#2E7D32"))

    # -- Exportacion ------------------------------------------------

    def _exportar_excel(self) -> None:
        registros = self._obtener_registros_filtrados()
        if not registros:
            QMessageBox.information(
                self, "Sin datos", "No hay registros para exportar."
            )
            return
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar como Excel",
            f"bitacora_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Archivos Excel (*.xlsx)",
        )
        if not ruta:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Bitacora"
            headers = ["Fecha", "Proceso", "Nivel", "Modo", "Modulo", "Mensaje"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill(
                    start_color="1976D2", end_color="1976D2", fill_type="solid"
                )
                c.font = Font(bold=True, color="FFFFFF")
                c.alignment = Alignment(horizontal="center")
            for r in registros:
                prueba = es_modo_prueba(r["mensaje"])
                modo_str = "PRUEBA" if prueba is True else (
                    "PROD" if prueba is False else ""
                )
                ws.append([
                    r["fecha"],
                    r.get("_proceso", ""),
                    r["nivel"],
                    modo_str,
                    r["modulo"],
                    quitar_marca_prueba(r["mensaje"]),
                ])
            # Ancho de columnas.
            for col, ancho in enumerate([20, 14, 10, 10, 12, 80]):
                ws.column_dimensions[chr(65 + col)].width = ancho
            wb.save(ruta)
            log().info("Bitacora exportada a Excel: %s", ruta)
            QMessageBox.information(
                self, "Exportado", f"Bitacora exportada a:\n{ruta}"
            )
        except Exception as e:
            log().exception("Error exportando a Excel: %s", e)
            QMessageBox.critical(
                self, "Error", f"No se pudo exportar:\n{e}"
            )

    def _exportar_csv(self) -> None:
        registros = self._obtener_registros_filtrados()
        if not registros:
            QMessageBox.information(
                self, "Sin datos", "No hay registros para exportar."
            )
            return
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar como CSV",
            f"bitacora_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "Archivos CSV (*.csv)",
        )
        if not ruta:
            return
        try:
            with open(ruta, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["Fecha", "Proceso", "Nivel", "Modo", "Modulo", "Mensaje"])
                for r in registros:
                    prueba = es_modo_prueba(r["mensaje"])
                    modo_str = "PRUEBA" if prueba is True else (
                        "PROD" if prueba is False else ""
                    )
                    w.writerow([
                        r["fecha"],
                        r.get("_proceso", ""),
                        r["nivel"],
                        modo_str,
                        r["modulo"],
                        quitar_marca_prueba(r["mensaje"]),
                    ])
            log().info("Bitacora exportada a CSV: %s", ruta)
            QMessageBox.information(
                self, "Exportado", f"Bitacora exportada a:\n{ruta}"
            )
        except Exception as e:
            log().exception("Error exportando a CSV: %s", e)
            QMessageBox.critical(
                self, "Error", f"No se pudo exportar:\n{e}"
            )

    def _obtener_registros_filtrados(self) -> list[dict]:
        """Repite el filtrado actual para exportacion."""
        fecha_desde = self._fecha_desde.date().toString("yyyy-MM-dd")
        fecha_hasta = self._fecha_hasta.date().toString("yyyy-MM-dd")
        proc = self._filtro_proceso.currentText()
        nivel = self._filtro_nivel.currentText()
        modo = self._filtro_modo.currentText()
        texto = self._buscador.text().strip().lower()

        out = []
        for r in self._registros:
            f = r["fecha"][:10]
            if f < fecha_desde or f > fecha_hasta:
                continue
            if nivel != "Todos" and r["nivel"] != nivel:
                continue
            p = _extraer_proceso(r["mensaje"]).capitalize()
            if proc != "Todos" and p != proc:
                continue
            if modo != "Todos":
                prueba = es_modo_prueba(r["mensaje"])
                if modo == "Prueba" and prueba is not True:
                    continue
                if modo == "Produccion" and prueba is not False:
                    continue
            if texto and texto not in r["mensaje"].lower():
                continue
            out.append({**r, "_proceso": p})
        return out

    # -- Limpieza de registros antiguos -----------------------------

    def _limpiar_antiguos(self) -> None:
        if not BITACORA_LOG.exists():
            QMessageBox.information(
                self, "Sin bitacora", "El archivo de bitacora no existe."
            )
            return
        dias, ok = QInputDialog.getInt(
            self,
            "Limpiar registros antiguos",
            "Eliminar registros con mas de N dias de antiguedad:",
            30, 1, 3650, 1,
        )
        if not ok:
            return
        resp = QMessageBox.warning(
            self,
            "Confirmar limpieza",
            f"Vas a eliminar todos los registros con mas de {dias} dias.\n"
            f"Esta accion NO se puede deshacer.\n\n¿Continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if resp != QMessageBox.StandardButton.Yes:
            return

        try:
            from datetime import timedelta
            cutoff = datetime.now() - timedelta(days=dias)
            kept: list[str] = []
            removed = 0
            with BITACORA_LOG.open("r", encoding="utf-8", errors="replace") as f:
                lineas = f.readlines()
            for ln in lineas:
                m = re.match(
                    r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", ln
                )
                if not m:
                    kept.append(ln)
                    continue
                try:
                    fecha = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    kept.append(ln)
                    continue
                if fecha >= cutoff:
                    kept.append(ln)
                else:
                    removed += 1
            with BITACORA_LOG.open("w", encoding="utf-8") as f:
                f.writelines(kept)
            log().info(
                "Bitacora limpiada: %d registros eliminados (>%d dias)",
                removed, dias,
            )
            QMessageBox.information(
                self,
                "Limpieza completada",
                f"Se eliminaron {removed} registros con mas de {dias} dias.",
            )
            self.refrescar()
        except Exception as e:
            log().exception("Error limpiando bitacora: %s", e)
            QMessageBox.critical(
                self, "Error", f"No se pudo limpiar la bitacora:\n{e}"
            )