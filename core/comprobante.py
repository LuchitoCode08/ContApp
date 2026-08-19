"""Proceso: Generar Comprobante Bancolombia.

Entrada: uno o más ZIPs con CSVs adentro (los CSV de Bancolombia tienen
una sola columna con valores separados por coma). Los CSVs de todos los
ZIPs se concatenan en un solo DataFrame.
Salida: 1 Excel con 5 hojas (Original, Por cuentas, Por conceptos,
        Intereses, Gastos bancarios) + 1 archivo FOAPAL (fzrcoco.xlsx).
"""
from __future__ import annotations

import zipfile
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Callable

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

from app.config import JSONS_DIR, RESULTADOS_DIR
from core.archivos import carpeta_modo_prueba, carpeta_resultados
from core.base import ProcesoBase, ProcesoCancelado, ResultadoProceso
from core.json_manager import leer_json

ENCODING_FALLBACKS: tuple[str, ...] = ("utf-8", "utf-8-sig", "latin-1")

# Estructura de las columnas del CSV de Bancolombia
COL_CUENTA = 0
COL_PREFIJO = 1
COL_TIPO = 2
COL_FECHA = 3
COL_ID = 4
COL_VALOR = 5
COL_CODIGO_CONCEPTO = 6
COL_CONCEPTO = 7
COL_VALOR_NUM = 8
COL_CODIGO_CONTABLE = 9

COLUMNAS_NUMERICAS: list[int] = [
    COL_CUENTA, COL_PREFIJO, COL_TIPO, COL_VALOR, COL_VALOR_NUM,
]
CODIGO_EXCEPCION_FOAPAL: str = "119090"
EXCEPCION_FOAPAL: dict = {
    "Fondo": "FOESPC",
    "Organizacion": 52617,
    "Cuenta": 280517,
    "Programa": 52617,
}
ENCABEZADOS_FOAPAL: list[str] = [
    "Codigo", "Concepto", "Valor", "Fecha",
    "Fondo", "Organizacion", "Cuenta", "Programa", "D/C",
]

MESES_ES: dict[int, str] = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _read_csv_con_fallback(archivo: Path) -> pd.DataFrame:
    """Lee un CSV desde disco con separador pipe para evitar partir comas en descripciones."""
    for encoding in ENCODING_FALLBACKS:
        try:
            return pd.read_csv(
                archivo, decimal=".", sep="|", header=None, dtype=str,
                encoding=encoding,
            )
        except UnicodeDecodeError:
            continue
    return pd.read_csv(
        archivo, decimal=".", sep="|", header=None, dtype=str, encoding="latin-1",
    )


class ProcesoComprobante(ProcesoBase):
    """Genera el comprobante contable y el archivo FOAPAL."""

    LOG_PREFIX = "[Comprobante]"

    def __init__(self) -> None:
        super().__init__()
        json_dir = JSONS_DIR / "comprobante"
        self.clasificador_conceptos: dict = leer_json(json_dir / "codigos_conceptos.json")
        self.codigos_contables: dict[str, str] = leer_json(json_dir / "codigos_contables.json")
        self.nit_bancolombia: dict[str, str] = leer_json(json_dir / "nit_bancolombia.json")
        self.foapal_config: dict = leer_json(json_dir / "foapal.json")
        self.codigos_ignorados: dict = leer_json(json_dir / "codigos_ignorados.json")

        self.cuenta_bancolombia: str = next(
            cuenta for cuenta, nit in self.nit_bancolombia.items()
            if nit != "890903938"
        )
        self.concepto_fiduciaria: str = next(
            codigo for codigo, descripciones
            in self.clasificador_conceptos["Gastos bancarios"].items()
            if any("PAGP A PROVE FIDUCIARIA BANC" in d for d in descripciones)
        )
        self.codigos_interes: list[str] = list(
            self.clasificador_conceptos.get("Intereses", {}).keys()
        )
        self.codigos_gastos: list[str] = list(
            self.clasificador_conceptos.get("Gastos bancarios", {}).keys()
        )

    @property
    def nombre(self) -> str:
        return "comprobante"

    @property
    def descripcion(self) -> str:
        return (
            "Genera el comprobante contable Bancolombia (Excel 5 hojas) "
            "y el archivo FOAPAL (fzrcoco.xlsx) a partir de un ZIP con CSVs."
        )

    @property
    def extensiones_entrada(self) -> tuple[str, ...]:
        return (".zip",)

    @property
    def extensiones_salida(self) -> tuple[str, ...]:
        return (".xlsx",)

    def validar_archivos(self, archivos: list[Path]) -> str | None:
        if not archivos:
            return "Se esperaba al menos un archivo ZIP."
        for archivo in archivos:
            if archivo.suffix.lower() not in self.extensiones_entrada:
                return (
                    f"El archivo '{archivo.name}' no es un ZIP. "
                    "Para comprobante solo se aceptan archivos .zip."
                )
            if not zipfile.is_zipfile(archivo):
                return f"El archivo '{archivo.name}' no es un ZIP válido."
        return None

    def _leer_csvs_de_zip(self, zip_path: Path) -> pd.DataFrame:
        """Lee todos los CSVs dentro del ZIP y devuelve un DataFrame único."""
        datos: list[pd.DataFrame] = []
        with zipfile.ZipFile(zip_path, "r") as zf:
            for nombre in zf.namelist():
                if not nombre.lower().endswith(".csv"):
                    continue
                with zf.open(nombre) as f:
                    raw = f.read()
                    df = self._read_csv_bytes(raw)
                datos.append(df)
        if not datos:
            return pd.DataFrame()
        return pd.concat(datos, ignore_index=True)

    def verificar_codigos_conceptos(self, df: pd.DataFrame) -> list[str]:
        """Devuelve códigos de concepto del CSV que no estén mapeados."""
        if df.empty or COL_CODIGO_CONCEPTO not in df.columns:
            return []

        codigos_csv = set(df[COL_CODIGO_CONCEPTO].astype(str).str.strip())
        codigos_csv.discard("")
        codigos_mapeados = set()
        codigos_mapeados.update(self.clasificador_conceptos.get("Intereses", {}).keys())
        codigos_mapeados.update(self.clasificador_conceptos.get("Gastos bancarios", {}).keys())
        codigos_mapeados.update(self.foapal_config.get("creditos", {}).keys())
        codigos_mapeados.update(self.foapal_config.get("debitos", {}).keys())
        codigos_mapeados.update(self.codigos_ignorados.get("codigos", []))

        codigos_no_encontrados = codigos_csv - codigos_mapeados
        return sorted(codigos_no_encontrados, key=lambda c: (len(c), c))

    def _split_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or 0 not in df.columns:
            return pd.DataFrame()
        splitted = df[0].str.split(pat=",", expand=True)
        return splitted.apply(lambda c: c.astype(str).str.strip())

    def obtener_codigos_desconocidos(
        self, archivos: list[Path],
    ) -> tuple[list[str], dict[str, str]]:
        """Lee todos los ZIPs y devuelve los códigos desconocidos + descripciones."""
        copias: list[pd.DataFrame] = []
        for zip_path in archivos:
            copias.append(self._leer_csvs_de_zip(zip_path))
        crudo = pd.concat(copias, ignore_index=True) if copias else pd.DataFrame()
        df = self._split_dataframe(crudo)

        codigos = self.verificar_codigos_conceptos(df)
        descripciones: dict[str, str] = {}
        if codigos and COL_CONCEPTO in df.columns:
            for codigo in codigos:
                filas = df[df[COL_CODIGO_CONCEPTO].astype(str).str.strip() == codigo]
                if not filas.empty:
                    descripciones[codigo] = str(filas.iloc[0][COL_CONCEPTO]).strip()
        return codigos, descripciones

    @staticmethod
    def _read_csv_bytes(raw: bytes) -> pd.DataFrame:
        for encoding in ENCODING_FALLBACKS:
            try:
                return pd.read_csv(
                    StringIO(raw.decode(encoding)),
                    decimal=".", sep="|", header=None, dtype=str,
                )
            except UnicodeDecodeError:
                continue
        return pd.read_csv(
            StringIO(raw.decode("latin-1")),
            decimal=".", sep="|", header=None, dtype=str,
        )

    def _agregar_codigo_contable(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        serie = df[0].astype(str).str.strip()
        nueva = serie.map(self.codigos_contables).fillna("No encontrado")
        df.insert(len(df.columns), "Codigo Contable", nueva)
        return df

    def _copy_data(
        self, copia: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        if copia.empty:
            empty = pd.DataFrame()
            return empty, empty, empty, empty

        por_cuentas = copia[0].str.split(pat=",", expand=True)
        por_cuentas = por_cuentas.apply(lambda c: c.astype(str).str.strip())

        mask_9729 = por_cuentas[COL_CODIGO_CONCEPTO] == self.concepto_fiduciaria
        mask_cuenta_valida = por_cuentas.eq(self.cuenta_bancolombia).any(axis=1)
        por_cuentas = por_cuentas[~mask_9729 | mask_cuenta_valida]

        por_cuentas[COL_FECHA] = pd.to_datetime(
            por_cuentas[COL_FECHA], format="%d%m%Y", errors="coerce"
        ).dt.strftime("%d/%m/%Y")

        for col in COLUMNAS_NUMERICAS:
            por_cuentas[col] = pd.to_numeric(por_cuentas[col], errors="coerce")

        por_cuentas[COL_CODIGO_CONCEPTO] = (
            por_cuentas[COL_CODIGO_CONCEPTO].astype(str).str.strip()
        )
        intereses = por_cuentas[
            por_cuentas[COL_CODIGO_CONCEPTO].isin(self.codigos_interes)
        ].copy()
        gastos = por_cuentas[
            por_cuentas[COL_CODIGO_CONCEPTO].isin(self.codigos_gastos)
        ].copy()

        por_cuentas = self._agregar_codigo_contable(por_cuentas)
        intereses = self._agregar_codigo_contable(intereses)
        gastos = self._agregar_codigo_contable(gastos)

        por_cuentas = por_cuentas.sort_values(by=[0], ascending=True)
        por_conceptos = por_cuentas.sort_values(by=[7], ascending=True)
        intereses = intereses.sort_values(by=[7], ascending=True)
        gastos = gastos.sort_values(by=[7], ascending=True)
        return por_cuentas, por_conceptos, intereses, gastos

    def _writer_excel(
        self, hojas: dict[str, pd.DataFrame], carpeta: Path,
    ) -> Path:
        fecha_actual = datetime.now()
        fecha_formateada = fecha_actual - relativedelta(months=1)
        mes_nombre = (
            f"{fecha_formateada.month:02d} "
            f"{MESES_ES[fecha_formateada.month].upper()} "
            f"{fecha_formateada.year}"
        )
        archivo_final = f"{mes_nombre} Bancolombia.xlsx"
        ruta_final = carpeta / archivo_final

        with pd.ExcelWriter(ruta_final, engine="openpyxl") as writer:
            for nombre_hoja, datos in hojas.items():
                datos.to_excel(
                    writer, sheet_name=nombre_hoja, header=False, index=False,
                )
                ws = writer.sheets[nombre_hoja]
                COL_VALOR_EN_EXCEL = COL_VALOR + 1
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=COL_VALOR_EN_EXCEL).number_format = "0.00"
                for col in ws.columns:
                    max_length = 0
                    column_letter = col[0].column_letter
                    column_index = col[0].column
                    for cell in col:
                        if column_index == COL_VALOR_EN_EXCEL and isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00; [Red]-0.00"
                        if cell.value:
                            largo = len(str(cell.value))
                            if largo > max_length:
                                max_length = largo
                    ws.column_dimensions[column_letter].width = max(max_length + 2, 10)
        return ruta_final

    def _codigo_para_foapal(self, cuenta: str) -> str:
        if not isinstance(cuenta, str):
            cuenta = str(cuenta).strip()
        return self.nit_bancolombia.get(cuenta, cuenta)

    def _aplicar_foapal(
        self, df_conceptos: pd.DataFrame, carpeta: Path,
        *, modo_prueba: bool = False,
        progreso: Callable[[int, int], None] | None = None,
        cancelado: Callable[[], bool] | None = None,
    ) -> Path | None:
        if df_conceptos.empty:
            return None
        try:
            fechas = pd.to_datetime(
                df_conceptos[COL_FECHA], format="%d/%m/%Y", errors="coerce"
            )
            df = df_conceptos.copy()
            df[COL_FECHA] = fechas.apply(
                lambda d: f"{d.day:02d}-{MESES_ES[d.month]}-{d.year}" if pd.notna(d) else ""
            )

            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet("Comprobante")
            for col_idx, encabezado in enumerate(ENCABEZADOS_FOAPAL, start=1):
                ws.cell(row=1, column=col_idx, value=encabezado)

            creditos = self.foapal_config.get("creditos", {})
            debitos = self.foapal_config.get("debitos", {})
            filas: list[list] = []
            total = len(df)
            paso = max(1, total // 100)

            if cancelado and total > 0 and cancelado():
                raise ProcesoCancelado()

            for i, (_, fila) in enumerate(df.iterrows()):
                if cancelado and i > 0 and i % 100 == 0 and cancelado():
                    raise ProcesoCancelado()
                if progreso and i > 0 and i % paso == 0:
                    progreso(i, total)
                clave_concepto = str(fila.iloc[COL_CODIGO_CONCEPTO]).strip()
                foapal_info = creditos.get(clave_concepto) or debitos.get(clave_concepto)
                if foapal_info is None:
                    continue
                codigo_contable = str(fila.iloc[COL_CODIGO_CONTABLE]).strip()
                concepto = fila.iloc[COL_CONCEPTO]
                valor = abs(fila.iloc[COL_VALOR])
                fecha = fila.iloc[COL_FECHA]

                codigo_universidad = self._codigo_para_foapal(
                    str(fila.iloc[COL_CUENTA]).strip()
                )

                try:
                    codigo_contable_num = int(codigo_contable)
                except ValueError:
                    codigo_contable_num = codigo_contable

                dc_principal = foapal_info.get("D/C", "C")
                dc_contraparte = "D" if dc_principal == "C" else "C"
                fondo = foapal_info.get("Fondo", "FOPNAL")
                organizacion = foapal_info.get("Organizacion", "999999")
                programa = foapal_info.get("Programa", "999999")

                filas.append([
                    codigo_universidad, concepto, valor, fecha,
                    fondo, organizacion, codigo_contable_num, programa, dc_principal,
                ])
                if codigo_contable == CODIGO_EXCEPCION_FOAPAL:
                    filas.append([
                        codigo_universidad, concepto, valor, fecha,
                        EXCEPCION_FOAPAL["Fondo"],
                        EXCEPCION_FOAPAL["Organizacion"],
                        EXCEPCION_FOAPAL["Cuenta"],
                        EXCEPCION_FOAPAL["Programa"],
                        dc_contraparte,
                    ])
                else:
                    filas.append([
                        codigo_universidad, concepto, valor, fecha,
                        fondo, organizacion, codigo_contable_num, programa, dc_contraparte,
                    ])

            for fila_idx, row_data in enumerate(filas, start=2):
                for col_idx, valor_celda in enumerate(row_data, start=1):
                    ws.cell(row=fila_idx, column=col_idx, value=valor_celda)
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        largo = len(str(cell.value))
                        if largo > max_length:
                            max_length = largo
                ws.column_dimensions[column_letter].width = max(max_length + 2, 10)

            ruta_foapal = carpeta / "fzrcoco.xlsx"
            wb.save(ruta_foapal)
            return ruta_foapal
        except ProcesoCancelado:
            raise
        except Exception:
            return None

    def ejecutar(
        self,
        archivos: list[Path],
        modo_prueba: bool = False,
        *,
        progreso: Callable[[int, int], None] | None = None,
        cancelado: Callable[[], bool] | None = None,
    ) -> ResultadoProceso:
        copias: list[pd.DataFrame] = []
        try:
            for zip_path in archivos:
                copias.append(self._leer_csvs_de_zip(zip_path))
        except Exception as e:
            return ResultadoProceso(exito=False, mensaje=f"No se pudo leer el ZIP: {e}")
        copia = pd.concat(copias, ignore_index=True) if copias else pd.DataFrame()

        por_cuentas, por_conceptos, intereses, gastos = self._copy_data(copia)

        if modo_prueba:
            carpeta = carpeta_modo_prueba(RESULTADOS_DIR, self.nombre)
        else:
            carpeta = carpeta_resultados(RESULTADOS_DIR, self.nombre)

        try:
            ruta_excel = self._writer_excel({
                "Original": copia,
                "Por cuentas": por_cuentas,
                "Por conceptos": por_conceptos,
                "Intereses": intereses,
                "Gastos bancarios": gastos,
            }, carpeta)
        except Exception as e:
            return ResultadoProceso(exito=False, mensaje=f"No se pudo escribir el Excel: {e}")

        ruta_foapal = self._aplicar_foapal(
            por_conceptos, carpeta, modo_prueba=modo_prueba,
            progreso=progreso, cancelado=cancelado,
        )

        archivos_generados: list[Path] = [ruta_excel]
        if ruta_foapal is not None:
            archivos_generados.append(ruta_foapal)

        return ResultadoProceso(
            exito=True,
            mensaje="Comprobante generado correctamente.",
            archivos_salida=archivos_generados,
            detalles={"filas_origen": len(copia)},
        )
